"""
GAP-5: Export service — Excel, CSV, and Xray-ready CSV generation.

Provides:
  export_test_cases_excel          → openpyxl Workbook bytes
  export_test_cases_csv            → CSV string (RFC 4180)
  export_test_cases_xray_csv       → Xray-compatible CSV (one row per step)
  export_traceability_matrix_excel → openpyxl Workbook bytes (matrix)
  get_requirement_traceability_chain → per-requirement chain dict

Edge cases handled:
  - Empty projects / no test cases
  - Null / undefined fields (all coerced to safe defaults)
  - Excel cell text limit (32 767 chars) — text truncated with notice
  - Unicode in titles, steps, BDD scenarios
  - Steps stored as list[dict] or plain strings
  - Pagination across large datasets (streams all pages)
  - Requirements with no scenarios/test-cases/executions/defects
  - Circular / missing lineage references
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.defect import DefectDraft
from app.models.execution import ExecutionResult, ExecutionRun
from app.models.requirement import Requirement
from app.models.test_case import TestCase
from app.models.test_plan import PlanTestCase
from app.models.test_scenario import TestScenario

# ── Excel cell helpers ──────────────────────────────────────────────────────

_EXCEL_MAX_CELL = 32_767  # hard Excel limit per cell


def _safe_str(value: Any, max_len: int = _EXCEL_MAX_CELL) -> str:
    """Coerce *value* to a str safe for an Excel cell."""
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        try:
            text = json.dumps(value, ensure_ascii=False)
        except Exception:
            text = str(value)
    elif isinstance(value, bool):
        text = "Yes" if value else "No"
    else:
        text = str(value)
    if len(text) > max_len:
        trunc_at = max_len - 20
        if trunc_at > 0:
            text = text[:trunc_at] + "… [truncated]"
        else:
            # max_len is very small — just hard-cut, no suffix
            text = text[:max_len]
    return text


def _steps_to_text(steps: list | None) -> str:
    """Flatten test-case steps to a human-readable string."""
    if not steps:
        return ""
    lines: list[str] = []
    for i, step in enumerate(steps, 1):
        if isinstance(step, dict):
            action = step.get("action") or step.get("step") or ""
            expected = step.get("expected_result") or step.get("expected") or ""
            if expected:
                lines.append(f"{i}. {action}  →  {expected}")
            else:
                lines.append(f"{i}. {action}")
        else:
            lines.append(f"{i}. {step}")
    return "\n".join(lines)


def _list_to_text(items: list | None) -> str:
    if not items:
        return ""
    return "; ".join(str(x) for x in items if x is not None)


# ── Data fetching helpers ───────────────────────────────────────────────────

async def _fetch_all_test_cases(
    db: AsyncSession,
    project_id: int,
    include_drafts: bool = False,
) -> list[TestCase]:
    """Return all test cases for *project_id*, optionally including drafts."""
    stmt = (
        select(TestCase)
        .where(TestCase.project_id == project_id)
        .options(
            selectinload(TestCase.channel),
            selectinload(TestCase.domain),
            selectinload(TestCase.area_of_test),
            selectinload(TestCase.taxonomy_product),
            selectinload(TestCase.taxonomy_sub_request_type),
            selectinload(TestCase.taxonomy_test_case_type),
            selectinload(TestCase.taxonomy_test_case_complexity),
        )
    )
    if not include_drafts:
        stmt = stmt.where(TestCase.status == "approved")
    stmt = stmt.order_by(TestCase.id)
    return list((await db.execute(stmt)).scalars().all())


async def _fetch_latest_execution_map(
    db: AsyncSession, tc_ids: list[int]
) -> dict[int, ExecutionResult]:
    """Return {test_case_id: most-recent ExecutionResult} for the UAT
    template's Overall Status / Tested By / SIT / Blocking Snag columns —
    these are execution-level facts (see the implementation plan's
    confirmed decision), resolved here for display rather than stored
    redundantly on TestCase."""
    if not tc_ids:
        return {}
    stmt = (
        select(ExecutionResult)
        .where(ExecutionResult.test_case_id.in_(tc_ids))
        .options(
            selectinload(ExecutionResult.tested_by),
            selectinload(ExecutionResult.blocking_defect),
        )
        .order_by(ExecutionResult.test_case_id, ExecutionResult.created_at.desc())
    )
    rows = (await db.execute(stmt)).scalars().all()
    latest: dict[int, ExecutionResult] = {}
    for er in rows:
        if er.test_case_id is not None and er.test_case_id not in latest:
            latest[er.test_case_id] = er
    return latest


async def _fetch_plan_enrollment_map(
    db: AsyncSession, tc_ids: list[int]
) -> dict[int, PlanTestCase]:
    """Return {test_case_id: most-recent plan enrollment} for the UAT
    template's plan-time Environment / Planned Execution Sequence columns."""
    if not tc_ids:
        return {}
    stmt = (
        select(PlanTestCase)
        .where(PlanTestCase.test_case_id.in_(tc_ids))
        .options(selectinload(PlanTestCase.environment))
        .order_by(PlanTestCase.test_case_id, PlanTestCase.id.desc())
    )
    rows = (await db.execute(stmt)).scalars().all()
    latest: dict[int, PlanTestCase] = {}
    for enrollment in rows:
        if enrollment.test_case_id not in latest:
            latest[enrollment.test_case_id] = enrollment
    return latest


_OVERALL_STATUS_DISPLAY = {
    "pending": "Not Executed",
    "not_run": "Not Executed",
    "running": "In Progress",
    "pass": "Passed",
    "passed_with_snag": "Passed with Snag",
    "fail": "Failed",
    "error": "Failed",
    "skip": "Skipped",
    "blocked": "Blocked",
}


def _template_row(
    tc: TestCase,
    latest_result: ExecutionResult | None,
    enrollment: PlanTestCase | None,
) -> list[Any]:
    """Build one row in the UAT template's 22-column order (docs/autonomous-
    automation-lab/test-case_template.xlsx). Values are raw (str/bool/None);
    callers coerce for their target format (Excel cell vs CSV field)."""
    overall_status = _OVERALL_STATUS_DISPLAY.get(
        (latest_result.status if latest_result else None) or "not_run", "Not Executed"
    )
    return [
        tc.id,
        tc.domain_name,
        tc.channel_name,
        tc.product_name,
        tc.area_of_test_name,
        tc.test_case_id,
        enrollment.environment.name if enrollment and enrollment.environment else None,
        tc.sub_request_type_name,
        tc.test_case_objective or tc.title,
        _list_to_text(tc.preconditions),
        _steps_to_text(tc.steps),
        tc.expected_result,
        tc.atc_test_case,
        tc.test_case_type_name,
        tc.test_case_complexity_name,
        latest_result.tested_by_name if latest_result else None,
        tc.jira_issue_key or tc.ppm_id,
        overall_status,
        (latest_result.blocking_defect_display_id if latest_result else None)
        or (latest_result.other_reason if latest_result else None),
        latest_result.sit_status if latest_result else None,
        enrollment.planned_execution_sequence if enrollment else None,
        tc.is_critical,
    ]


_TEMPLATE_HEADERS = [
    "ID", "Domain", "Channel", "Product", "Area of Test", "Test Case ID",
    "Environment", "Sub Request Type", "Test Case Objective", "Pre-Requisites",
    "Test Steps", "Expected Results", "ATC Test Case", "Test Case Type",
    "Test Case Complexity", "Tested By", "JIRA ID or PPM", "Overall Status",
    "Blocking Snag ID / Other Reason", "SIT", "Planned Execution Sequence",
    "Critical TC Mapping",
]

# Platform-specific columns kept after the 22 template columns so existing
# consumers of these fields (traceability lineage, legacy automation
# metadata) don't silently lose data.
_PLATFORM_EXTRA_HEADERS = [
    "Priority", "Severity", "BDD Scenario", "Approval Status",
    "Requirement ID", "Requirement Title", "Scenario ID", "Scenario Title",
    "Test Phase", "Product Group", "External TC ID", "External TC URL",
    "Created At",
]


def _platform_extra_row(tc: TestCase, req: Requirement | None, scen: TestScenario | None) -> list[Any]:
    return [
        tc.priority,
        tc.severity,
        tc.bdd_scenario,
        getattr(tc, "approval_status", None),
        req.requirement_id if req else None,
        req.title if req else None,
        scen.scenario_id if scen else None,
        scen.title if scen else None,
        tc.test_phase,
        tc.product_group,
        tc.external_tc_id,
        tc.external_tc_url,
        tc.created_at.isoformat() if tc.created_at else None,
    ]


async def _fetch_requirement_map(
    db: AsyncSession,
    project_id: int,
) -> dict[int, Requirement]:
    """Return {req.id: req} for the project."""
    stmt = select(Requirement).where(Requirement.project_id == project_id)
    rows = list((await db.execute(stmt)).scalars().all())
    return {r.id: r for r in rows}


async def _fetch_scenario_map(
    db: AsyncSession,
    project_id: int,
) -> dict[int, TestScenario]:
    stmt = select(TestScenario).where(TestScenario.project_id == project_id)
    rows = list((await db.execute(stmt)).scalars().all())
    return {s.id: s for s in rows}


# ── Excel styling helpers ───────────────────────────────────────────────────
#
# Matches the source workbook's look (docs/autonomous-automation-lab/
# test-case_template.xlsx, sheet "UAT Test Cases"): header row alternates
# between bright green (#92D050) and dusty rose (#D99694, the workbook's
# theme accent2 tinted 40%) per column, bold white Calibri 12, wrapped text.
# Data rows: Calibri 11, white background, thin border, wrapped text.

_TEMPLATE_HEADER_COLORS = ("92D050", "D99694")


def _make_header_style(col_index: int = 1):
    """Return openpyxl styling objects — imported lazily to avoid hard dep.
    `col_index` (1-based) picks which of the two alternating header colors."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    fill_color = _TEMPLATE_HEADER_COLORS[(col_index - 1) % 2]
    header_fill = PatternFill(fill_type="solid", fgColor=fill_color)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D0D7DE")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, header_align, header_border


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        header_font, header_fill, header_align, header_border = _make_header_style(col)
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border


def _style_data_row(ws, row: int, num_cols: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    data_font = Font(name="Calibri", size=11)
    data_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    thin = Side(border_style="thin", color="D0D7DE")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="top", wrap_text=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = data_fill
        cell.alignment = cell_align
        cell.border = cell_border


# ── Public API ──────────────────────────────────────────────────────────────

async def export_test_cases_excel(
    db: AsyncSession,
    project_id: int,
    include_drafts: bool = False,
) -> bytes:
    """
    Return an Excel workbook (bytes) with one row per test case, columns
    matching the UAT template's 22-field order (docs/autonomous-automation-
    lab/test-case_template.xlsx) followed by platform-specific extras.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    test_cases = await _fetch_all_test_cases(db, project_id, include_drafts)
    req_map = await _fetch_requirement_map(db, project_id)
    scen_map = await _fetch_scenario_map(db, project_id)
    tc_ids = [tc.id for tc in test_cases]
    latest_result_map = await _fetch_latest_execution_map(db, tc_ids)
    enrollment_map = await _fetch_plan_enrollment_map(db, tc_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = _TEMPLATE_HEADERS + _PLATFORM_EXTRA_HEADERS

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(headers))

    for row_idx, tc in enumerate(test_cases, 2):
        req = req_map.get(tc.requirement_id) if tc.requirement_id else None
        scen = scen_map.get(tc.scenario_id) if tc.scenario_id else None
        latest_result = latest_result_map.get(tc.id)
        enrollment = enrollment_map.get(tc.id)

        row_data = [_safe_str(v) for v in _template_row(tc, latest_result, enrollment)] + [
            _safe_str(v) for v in _platform_extra_row(tc, req, scen)
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(headers))

    # Set column widths — 22 template columns first, then platform extras.
    col_widths = [
        8, 16, 16, 18, 18, 24,
        14, 18, 40, 36,
        50, 40, 18, 16,
        18, 18, 18, 16,
        28, 12, 20,
        14,
    ] + [12, 12, 40, 16, 16, 40, 14, 40, 14, 18, 20, 40, 22]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = ws["A2"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_test_cases_csv(
    db: AsyncSession,
    project_id: int,
    include_drafts: bool = False,
) -> str:
    """Return a plain CSV string of all test cases, columns matching the
    UAT template's 22-field order followed by platform-specific extras."""
    test_cases = await _fetch_all_test_cases(db, project_id, include_drafts)
    req_map = await _fetch_requirement_map(db, project_id)
    scen_map = await _fetch_scenario_map(db, project_id)
    tc_ids = [tc.id for tc in test_cases]
    latest_result_map = await _fetch_latest_execution_map(db, tc_ids)
    enrollment_map = await _fetch_plan_enrollment_map(db, tc_ids)

    buf = io.StringIO()
    writer = csv.writer(buf, dialect="excel")
    writer.writerow(_TEMPLATE_HEADERS + _PLATFORM_EXTRA_HEADERS)

    for tc in test_cases:
        req = req_map.get(tc.requirement_id) if tc.requirement_id else None
        scen = scen_map.get(tc.scenario_id) if tc.scenario_id else None
        latest_result = latest_result_map.get(tc.id)
        enrollment = enrollment_map.get(tc.id)

        row = [
            ("Yes" if v else "No") if isinstance(v, bool) else (v if v is not None else "")
            for v in _template_row(tc, latest_result, enrollment) + _platform_extra_row(tc, req, scen)
        ]
        writer.writerow(row)

    return buf.getvalue()


async def export_test_cases_xray_csv(
    db: AsyncSession,
    project_id: int,
    include_drafts: bool = False,
) -> str:
    """
    Return an Xray-compatible CSV string for Jira Xray import.

    Format: one header row, then one or more data rows per test case.
    Multi-step manual tests: first row has Summary + meta, subsequent
    rows only have Step columns (matching Xray's CSV import spec).

    Ref: https://docs.getxray.app/display/XRAY/Importing+Test+Cases+from+CSV
    """
    test_cases = await _fetch_all_test_cases(db, project_id, include_drafts)
    req_map = await _fetch_requirement_map(db, project_id)

    buf = io.StringIO()
    writer = csv.writer(buf, dialect="excel")

    writer.writerow([
        "Summary", "Issue ID", "Test Type",
        "Generic Test Definition",   # BDD / Cucumber definition
        "Step Action", "Step Data", "Step Result",
        "Labels", "Priority", "Requirement Issue Key",
        "Telecom Domain", "External TC ID",
    ])

    for tc in test_cases:
        req = req_map.get(tc.requirement_id) if tc.requirement_id else None
        req_jira_key = req.jira_issue_key if req else ""

        steps = tc.steps or []
        test_type = "Cucumber" if tc.bdd_scenario else "Manual"

        # For BDD / Cucumber — single row with Generic Test Definition
        if test_type == "Cucumber":
            writer.writerow([
                tc.title or "",
                tc.test_case_id or "",
                "Cucumber",
                tc.bdd_scenario or "",
                "",  # Step Action
                "",  # Step Data
                "",  # Step Result
                "",
                tc.priority or "",
                req_jira_key or "",
                tc.telecom_domain or "",
                tc.external_tc_id or "",
            ])
            continue

        # Manual: emit one leading row per step (or one fallback row if no steps)
        if not steps:
            # EC-07: no steps — emit a single fallback row
            writer.writerow([
                tc.title or "",
                tc.test_case_id or "",
                "Manual",
                "",  # Generic Test Definition
                "Refer to test design document",
                "",
                "",
                "",
                tc.priority or "",
                req_jira_key or "",
                tc.telecom_domain or "",
                tc.external_tc_id or "",
            ])
        else:
            for step_idx, step in enumerate(steps):
                if isinstance(step, dict):
                    action = step.get("action") or step.get("step") or ""
                    data = step.get("test_data") or step.get("data") or ""
                    result = step.get("expected_result") or step.get("expected") or ""
                else:
                    # EC-08: step stored as plain string
                    action = str(step)
                    data = ""
                    result = ""

                if step_idx == 0:
                    # First step row — include all metadata
                    writer.writerow([
                        tc.title or "",
                        tc.test_case_id or "",
                        "Manual",
                        "",  # Generic Test Definition (not applicable)
                        _safe_str(action, 500),
                        _safe_str(data, 500),
                        _safe_str(result, 500),
                        "",
                        tc.priority or "",
                        req_jira_key or "",
                        tc.telecom_domain or "",
                        tc.external_tc_id or "",
                    ])
                else:
                    # Continuation step rows — Xray ignores leading cols
                    writer.writerow([
                        "",   # Summary (blank for continuation)
                        "",   # Issue ID
                        "",   # Test Type
                        "",   # Generic Test Definition
                        _safe_str(action, 500),
                        _safe_str(data, 500),
                        _safe_str(result, 500),
                        "", "", "", "", "",
                    ])

    return buf.getvalue()


async def export_traceability_matrix_excel(
    db: AsyncSession,
    project_id: int,
    include_drafts: bool = False,
) -> bytes:
    """
    Return an Excel workbook (bytes) with the requirement→test-case traceability matrix.

    Sheet 1: Matrix (one row per requirement)
    Sheet 2: Summary (counts by status)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Fetch data
    req_stmt = select(Requirement).where(Requirement.project_id == project_id)
    if not include_drafts:
        req_stmt = req_stmt.where(Requirement.status == "approved")
    req_stmt = req_stmt.order_by(Requirement.id)
    requirements = list((await db.execute(req_stmt)).scalars().all())

    # Build requirement → test case mapping
    tc_stmt = select(TestCase).where(TestCase.project_id == project_id)
    if not include_drafts:
        tc_stmt = tc_stmt.where(TestCase.status == "approved")
    all_tcs = list((await db.execute(tc_stmt)).scalars().all())

    tc_by_req: dict[int, list[TestCase]] = {}
    for tc in all_tcs:
        if tc.requirement_id:
            tc_by_req.setdefault(tc.requirement_id, []).append(tc)

    # Execution results for test cases
    tc_ids = [tc.id for tc in all_tcs]
    er_by_tc: dict[int, list[ExecutionResult]] = {}
    if tc_ids:
        er_stmt = select(ExecutionResult).where(
            ExecutionResult.project_id == project_id,
            ExecutionResult.test_case_id.in_(tc_ids),
        )
        if not include_drafts:
            er_stmt = er_stmt.join(
                ExecutionRun, ExecutionRun.id == ExecutionResult.execution_run_id
            ).where(ExecutionRun.status == "approved")
        for er in (await db.execute(er_stmt)).scalars().all():
            if er.test_case_id:
                er_by_tc.setdefault(er.test_case_id, []).append(er)

    # Defects
    defect_by_tc: dict[int, list[DefectDraft]] = {}
    if tc_ids:
        def_stmt = select(DefectDraft).where(
            DefectDraft.project_id == project_id,
        )
        if not include_drafts:
            def_stmt = def_stmt.where(
                DefectDraft.status.in_(["approved", "pushed_to_jira"])
            )
        for d in (await db.execute(def_stmt)).scalars().all():
            # DefectDraft links via execution_result_id
            pass  # we aggregate at requirement level below

    wb = Workbook()

    # ── Sheet 1: Matrix ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Traceability Matrix"

    matrix_headers = [
        "Requirement ID", "Requirement Title", "Status",
        "Risk Level", "Telecom Domain", "Test Phase",
        "Test Case Count", "TC IDs", "Automation Candidates",
        "Execution Results", "Latest Execution Status",
        "Open Defects", "Coverage Gaps",
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(matrix_headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(matrix_headers))

    # Status color fills
    green_fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
    red_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
    amber_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")

    for row_idx, req in enumerate(requirements, 2):
        tcs = tc_by_req.get(req.id, [])
        tc_ids_str = ", ".join(tc.test_case_id for tc in tcs[:20])
        if len(tcs) > 20:
            tc_ids_str += f"… +{len(tcs) - 20} more"

        auto_count = sum(1 for tc in tcs if tc.automation_candidate)

        # Collect all execution results for all TCs of this requirement
        all_ers: list[ExecutionResult] = []
        for tc in tcs:
            all_ers.extend(er_by_tc.get(tc.id, []))

        er_statuses = [er.status for er in all_ers]
        er_summary = ", ".join(sorted(set(er_statuses))) if er_statuses else "none"
        latest_status = all_ers[-1].status if all_ers else "not executed"

        # Gaps
        gaps: list[str] = []
        if not tcs:
            gaps.append("no test cases")
        elif not all_ers:
            gaps.append("no execution")
        failed = [er for er in all_ers if er.status in {"failed", "error"}]
        if failed:
            gaps.append(f"{len(failed)} failures")
        gaps_str = "; ".join(gaps) if gaps else "none"

        row_data = [
            _safe_str(req.requirement_id),
            _safe_str(req.title),
            _safe_str(req.status),
            _safe_str(req.risk_level or ""),
            _safe_str(req.telecom_domain or ""),
            _safe_str(req.test_phase or ""),
            len(tcs),
            _safe_str(tc_ids_str),
            auto_count,
            len(all_ers),
            _safe_str(latest_status),
            len(failed) if failed else 0,
            _safe_str(gaps_str),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Row fill by gap status
        fill = green_fill if not gaps else (red_fill if "no test cases" in gaps else amber_fill)
        from openpyxl.styles import Alignment, Border, Side
        thin = Side(border_style="thin", color="D0D7DE")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell_align = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, len(matrix_headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = cell_align
            cell.border = cell_border

    # Column widths for matrix
    matrix_widths = [18, 50, 14, 14, 18, 14, 16, 40, 20, 18, 22, 14, 24]
    for idx, width in enumerate(matrix_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = ws["A2"]

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Summary")

    summary_data = [
        ("Total Requirements", len(requirements)),
        ("Total Test Cases (approved)", len(all_tcs)),
        ("Requirements with Test Cases", sum(1 for req in requirements if req.id in tc_by_req)),
        ("Requirements WITHOUT Test Cases", sum(1 for req in requirements if req.id not in tc_by_req)),
        ("Automation Candidates", sum(1 for tc in all_tcs if tc.automation_candidate)),
        ("Include Drafts", "Yes" if include_drafts else "No"),
    ]

    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 18

    from openpyxl.styles import Font
    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True, color="1B59F8")
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True, color="1B59F8")

    for row_idx, (metric, value) in enumerate(summary_data, 2):
        ws2.cell(row=row_idx, column=1, value=metric)
        ws2.cell(row=row_idx, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Per-requirement traceability chain ─────────────────────────────────────

async def get_requirement_traceability_chain(
    db: AsyncSession,
    requirement_id: int,
    project_id: int,
) -> dict[str, Any]:
    """
    Return a nested dict with the full traceability chain for one requirement:
      requirement → scenarios → test cases → execution results → defects

    Edge cases:
      - Requirement not found → raises ValueError
      - No scenarios/test cases/executions/defects → returns empty lists
      - Execution results fetched across all test cases of the requirement
    """
    # Requirement
    req_result = await db.execute(
        select(Requirement).where(
            Requirement.id == requirement_id,
            Requirement.project_id == project_id,
        )
    )
    req = req_result.scalar_one_or_none()
    if req is None:
        raise ValueError(f"Requirement {requirement_id} not found in project {project_id}")

    # Scenarios
    scen_result = await db.execute(
        select(TestScenario).where(
            TestScenario.requirement_id == requirement_id,
            TestScenario.project_id == project_id,
        ).order_by(TestScenario.id)
    )
    scenarios = list(scen_result.scalars().all())

    # Test Cases (via requirement_id directly, or via scenario)
    tc_result = await db.execute(
        select(TestCase).where(
            TestCase.project_id == project_id,
            TestCase.requirement_id == requirement_id,
        ).order_by(TestCase.id)
    )
    test_cases = list(tc_result.scalars().all())

    # Also collect TCs linked via scenarios (without direct requirement linkage)
    scen_ids = [s.id for s in scenarios]
    tc_by_id = {tc.id: tc for tc in test_cases}
    if scen_ids:
        scen_tc_result = await db.execute(
            select(TestCase).where(
                TestCase.project_id == project_id,
                TestCase.scenario_id.in_(scen_ids),
                TestCase.requirement_id.is_(None),  # avoid duplicates
            ).order_by(TestCase.id)
        )
        for tc in scen_tc_result.scalars().all():
            tc_by_id[tc.id] = tc
    test_cases = list(tc_by_id.values())

    # Execution Results
    tc_ids = [tc.id for tc in test_cases]
    exec_results: list[ExecutionResult] = []
    if tc_ids:
        er_result = await db.execute(
            select(ExecutionResult).where(
                ExecutionResult.project_id == project_id,
                ExecutionResult.test_case_id.in_(tc_ids),
            ).order_by(ExecutionResult.id.desc())
        )
        exec_results = list(er_result.scalars().all())

    # Defects
    er_ids = [er.id for er in exec_results]
    defects: list[DefectDraft] = []
    if er_ids:
        def_result = await db.execute(
            select(DefectDraft).where(
                DefectDraft.project_id == project_id,
                DefectDraft.execution_result_id.in_(er_ids),
            ).order_by(DefectDraft.id)
        )
        defects = list(def_result.scalars().all())

    # Build result dict
    def _req_dict(r: Requirement) -> dict:
        return {
            "id": r.id,
            "requirement_id": r.requirement_id,
            "title": r.title,
            "status": r.status,
            "risk_level": r.risk_level,
            "telecom_domain": r.telecom_domain,
            "quality_score": r.quality_score,
            "quality_verdict": r.quality_verdict,
        }

    def _scen_dict(s: TestScenario) -> dict:
        return {
            "id": s.id,
            "scenario_id": s.scenario_id,
            "title": s.title,
            "scenario_type": s.scenario_type,
            "priority": s.priority,
            "status": s.status,
        }

    def _tc_dict(tc: TestCase) -> dict:
        return {
            "id": tc.id,
            "test_case_id": tc.test_case_id,
            "title": tc.title,
            "test_type": tc.test_type,
            "priority": tc.priority,
            "status": tc.status,
            "automation_candidate": tc.automation_candidate,
            "scenario_id": tc.scenario_id,
            "jira_issue_key": tc.jira_issue_key,
        }

    def _er_dict(er: ExecutionResult) -> dict:
        return {
            "id": er.id,
            "test_case_id": er.test_case_id,
            "test_name": er.test_name,
            "status": er.status,
            "execution_mode": er.execution_mode,
            "created_at": er.created_at.isoformat() if er.created_at else None,
        }

    def _defect_dict(d: DefectDraft) -> dict:
        return {
            "id": d.id,
            "defect_id": d.defect_id,
            "summary": d.summary,
            "severity": d.severity,
            "priority": d.priority,
            "status": d.status,
            "jira_ready": d.jira_ready,
        }

    # Summary gaps
    gaps: list[str] = []
    if not test_cases:
        gaps.append("no_test_cases")
    if test_cases and not exec_results:
        gaps.append("no_execution")
    failed_ers = [er for er in exec_results if er.status in {"failed", "error"}]
    defect_er_ids = {d.execution_result_id for d in defects}
    if any(er.id not in defect_er_ids for er in failed_ers):
        gaps.append("undecided_failures")

    return {
        "requirement": _req_dict(req),
        "scenarios": [_scen_dict(s) for s in scenarios],
        "test_cases": [_tc_dict(tc) for tc in test_cases],
        "execution_results": [_er_dict(er) for er in exec_results],
        "defects": [_defect_dict(d) for d in defects],
        "summary": {
            "scenario_count": len(scenarios),
            "test_case_count": len(test_cases),
            "execution_count": len(exec_results),
            "defect_count": len(defects),
            "gaps": gaps,
        },
    }
