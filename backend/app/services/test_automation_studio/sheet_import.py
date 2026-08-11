"""Read a test case sheet in the platform's canonical template, exactly.

A test case document that already follows `test_case_template` is a table, not
prose. Sending it through the LLM extraction pass costs a call per segment,
caps what can be read at the model's output budget, and returns only the four
fields that pass asks for — so Domain, Channel, Product, Environment and every
execution column were lost before the export could reach them.

Parsing it directly costs nothing, cannot truncate, and keeps every column the
team filled in. The parsing helpers are the importer's own, so a sheet this
module reads and a sheet the platform imports are read the same way; anything
that does not match the template still goes to the agent.
"""
from __future__ import annotations

import io
import os
from typing import Any

from app.services.test_case_import_service import (
    _map_headers,
    _normalize_header,
    _parse_csv,
    _parse_excel,
    _parse_steps,
    _split_lines,
)
from app.services.test_case_template import TEST_CASE_TEMPLATE_HEADER_ALIASES

# Recognising the template needs more than one column in common: a stray sheet
# with an "ID" header is not this template, and reading it as one would produce
# rows with a title and nothing else.
MIN_TEMPLATE_FIELDS = 4
# Without these there is no test case to speak of, whatever else matched.
REQUIRED_TEMPLATE_FIELDS = {"test_case_id", "id"}

PARSEABLE_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # pandas renders an empty cell as NaN once it has seen a number in the
    # column; writing "NaN" into a user's spreadsheet would be worse than
    # leaving the cell empty.
    return "" if text.lower() in {"nan", "nat", "none"} else text


def can_parse(filename: str | None) -> bool:
    return bool(filename) and os.path.splitext(filename)[1].lower() in PARSEABLE_EXTENSIONS


def read_header_format(contents: bytes, filename: str) -> dict[str, dict]:
    """The header styling the uploaded workbook used, keyed by canonical field.

    The download is the same document handed back, so it should look like the
    one that was uploaded. The colours are not decoration here: this team tints
    each column by who owns it — green for the columns QA authors, salmon for
    the ones filled from the request, grey for execution results — so a
    uniformly styled download loses information a reader uses to scan the
    sheet.

    Returns {} for a CSV or an unreadable file; the caller then falls back to
    its own defaults rather than failing an export over cosmetics.
    """
    if os.path.splitext(filename)[1].lower() not in {".xlsx", ".xlsm"}:
        return {}
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(contents))
        sheet = workbook[workbook.sheetnames[0]]
    except Exception:
        return {}

    styles: dict[str, dict] = {}
    for cell in next(sheet.iter_rows(min_row=1, max_row=1), ()):
        header = str(cell.value or "").strip()
        field = TEST_CASE_TEMPLATE_HEADER_ALIASES.get(_normalize_header(header)) if header else None
        if not field:
            continue
        fill = cell.fill
        font = cell.font
        styles[field] = {
            # openpyxl reports theme/indexed colours as objects rather than
            # strings; only a literal ARGB can be replayed, so anything else is
            # dropped and the default is used for that column.
            "fill": fill.fgColor.rgb if fill and isinstance(getattr(fill.fgColor, "rgb", None), str) else None,
            "font_color": font.color.rgb if font and isinstance(getattr(font.color, "rgb", None), str) else None,
            "bold": bool(font.bold) if font else True,
            "size": float(font.sz) if font and font.sz else 12.0,
            "name": font.name if font and font.name else "Calibri",
            "width": (
                sheet.column_dimensions[cell.column_letter].width
                if cell.column_letter in sheet.column_dimensions
                else None
            ),
        }
    return styles


def parse_sheet(contents: bytes, filename: str) -> list[dict] | None:
    """Return one dict per test case row, or None if this is not the template.

    None means "not recognised, use the agent" — never an error. A document
    that does not follow the template is a normal thing to upload.
    """
    extension = os.path.splitext(filename)[1].lower()
    try:
        if extension == ".csv":
            headers, rows = _parse_csv(contents)
        else:
            headers, rows = _parse_excel(contents)
    except Exception:
        # Unreadable as a table is not a failure here: the agent gets its turn.
        return None

    mapping = _map_headers(headers)
    fields = set(mapping.values())
    if len(fields) < MIN_TEMPLATE_FIELDS or not (fields & REQUIRED_TEMPLATE_FIELDS):
        return None

    parsed: list[dict] = []
    for index, raw in enumerate(rows, start=1):
        row = {field: _clean(raw.get(header)) for header, field in mapping.items()}

        # The sheet may carry both an `ID` and a `Test Case ID`. The latter is
        # the team's identifier for the test case and wins; `ID` is a row
        # number that would make a meaningless display ID.
        display_id = row.get("test_case_id") or row.get("id")
        title = row.get("test_case_objective") or row.get("title")
        # No ID means this is not a test case row. A cell containing line
        # breaks makes the reader emit the continuation as its own row — the
        # tail of TC-13's expected result arrived as a row reading
        # ") remain unaffected." — and a template sheet always identifies a
        # real test case. Inventing an ID for one of these would put a phantom
        # test case in the workbench and, worse, preserve its made-up ID.
        if not display_id:
            continue
        if not title:
            title = display_id

        parsed.append(
            {
                "test_case_id": display_id,
                "title": title or display_id,
                "summary": row.get("test_case_objective") or None,
                "steps": _split_lines(row.get("steps")),
                "structured_steps": _parse_steps(row.get("steps")),
                "preconditions": _split_lines(row.get("preconditions")),
                "source_ref": f"{filename}, row {index}",
                # Everything the sheet had, for the download to hand back.
                "source_row": row,
            }
        )
    return parsed
