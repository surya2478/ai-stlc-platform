"""Screen 2 downloads — the same spreadsheet shape the team uploaded.

The download hands work back to the people who supplied it, so it uses the
platform's canonical test case template (`app/services/test_case_template.py`),
column for column and in order — the same contract the platform's own export
and its importer speak. A team can therefore round-trip: upload their sheet,
refine it here, download it, and open the result next to the original.

The studio owns a few of those columns and nothing else. It rewrites the
objective, pre-requisites, steps and expected result for automation, and it
decides classification; Domain, Channel, Product, Area of Test, Environment,
Sub Request Type and every execution column belong to the source and are
carried through untouched. Values are resolved in that order of authority:

    the refined test case  →  the row it came from  →  blank

Manual and Automation stay on separate sheets, because the two audiences act
on them separately — but both now carry the identical column layout, so
neither is a different document from the one that was uploaded.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.test_automation_studio import (
    TasDerivedRequirement,
    TasRefinedTestCase,
    TasSourceTestCase,
)
from app.models.document import UploadedDocument
from app.models.test_case import TestCase
from app.services.test_automation_studio import sheet_import
from app.services.test_case_template import (
    TEST_CASE_TEMPLATE_COLUMNS,
    TEST_CASE_TEMPLATE_HEADERS,
)

# Excel refuses a cell over 32767 characters, and a long generated step list
# will exceed it. Truncating is better than the whole export failing.
_EXCEL_MAX_CELL = 32000

# Appended after the template's own columns, never interleaved: a reader
# diffing this against their upload sees the familiar sheet first and the
# studio's additions after it, rather than their columns shifted sideways.
STUDIO_COLUMNS = [
    "Classification",
    "Classification Reason",
    "Application URL",
    "Step Targets",
    "Test Data Keys",
    "Test Data Required",
    "Test Data Notes",
    "Studio Source",
    "Studio Version",
]

COLUMNS = TEST_CASE_TEMPLATE_HEADERS + STUDIO_COLUMNS


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (dict, list)):
        rendered = json.dumps(value, ensure_ascii=False, default=str)
    else:
        rendered = str(value)
    return rendered[:_EXCEL_MAX_CELL]


def _steps_text(steps: Any) -> str:
    """Render steps the way the importer reads them back.

    `1. action -> expected` per line is the format `_parse_steps` parses, so a
    downloaded sheet can be re-imported without the steps collapsing into
    prose.
    """
    if not isinstance(steps, list):
        return _text(steps)
    lines: list[str] = []
    for step in steps:
        if not isinstance(step, dict):
            lines.append(str(step))
            continue
        number = step.get("step_number")
        action = step.get("action") or ""
        expected = step.get("expected_result")
        line = f"{number}. {action}" if number else str(action)
        if expected:
            line += f" -> {expected}"
        lines.append(line)
    return _text("\n".join(lines))


def _targets_text(steps: Any) -> str:
    if not isinstance(steps, list):
        return ""
    parts = [
        f"{step.get('step_number')}: {step.get('target')}"
        for step in steps
        if isinstance(step, dict) and step.get("target")
    ]
    return _text("\n".join(parts))


def _data_keys_text(test_case: TasRefinedTestCase) -> str:
    requirements = test_case.test_data_requirements or []
    if not isinstance(requirements, list):
        return ""
    parts = []
    for req in requirements:
        if not isinstance(req, dict):
            continue
        rendered = str(req.get("key"))
        if req.get("resolution"):
            rendered += f" [{req['resolution']}]"
        if req.get("example_value"):
            rendered += f" = {req['example_value']}"
        parts.append(rendered)
    return _text("\n".join(parts))


def _list_text(values: Any) -> str:
    if isinstance(values, list):
        return _text("\n".join(str(value) for value in values))
    return _text(values)


class _Context:
    """The rows a refined test case was built from, indexed for lookup."""

    def __init__(
        self,
        sources: dict[int, TasSourceTestCase],
        platform: dict[int, TestCase],
        requirements: dict[int, TasDerivedRequirement],
        header_format: dict[str, dict] | None = None,
    ) -> None:
        self.sources = sources
        self.platform = platform
        self.requirements = requirements
        # How the uploaded workbook styled its own header, keyed by canonical
        # field. Empty when the source was a CSV or is no longer on disk.
        self.header_format = header_format or {}

    def source_row(self, test_case: TasRefinedTestCase) -> dict:
        source = self.sources.get(test_case.source_uploaded_test_case_id or -1)
        row = getattr(source, "source_row", None)
        return row if isinstance(row, dict) else {}

    def platform_case(self, test_case: TasRefinedTestCase) -> TestCase | None:
        # Either link can reach the platform row: refined straight from it, or
        # refined from a sheet whose ID matched one.
        if test_case.source_test_case_id:
            return self.platform.get(test_case.source_test_case_id)
        source = self.sources.get(test_case.source_uploaded_test_case_id or -1)
        if source is not None and source.matched_platform_test_case_id:
            return self.platform.get(source.matched_platform_test_case_id)
        return None


async def build_context(db: AsyncSession, test_cases: list[TasRefinedTestCase]) -> _Context:
    """Load everything the rows inherit from, in three queries rather than 3N."""
    source_ids = {tc.source_uploaded_test_case_id for tc in test_cases if tc.source_uploaded_test_case_id}
    requirement_ids = {tc.derived_requirement_id for tc in test_cases if tc.derived_requirement_id}

    sources: dict[int, TasSourceTestCase] = {}
    if source_ids:
        rows = (
            await db.execute(select(TasSourceTestCase).where(TasSourceTestCase.id.in_(source_ids)))
        ).scalars().all()
        sources = {row.id: row for row in rows}

    platform_ids = {tc.source_test_case_id for tc in test_cases if tc.source_test_case_id}
    platform_ids |= {
        row.matched_platform_test_case_id for row in sources.values() if row.matched_platform_test_case_id
    }
    platform: dict[int, TestCase] = {}
    if platform_ids:
        rows = (
            await db.execute(
                select(TestCase)
                .where(TestCase.id.in_(platform_ids))
                .options(
                    selectinload(TestCase.channel),
                    selectinload(TestCase.domain),
                    selectinload(TestCase.area_of_test),
                    selectinload(TestCase.taxonomy_product),
                    selectinload(TestCase.taxonomy_sub_request_type),
                    selectinload(TestCase.taxonomy_test_case_type),
                    selectinload(TestCase.taxonomy_test_case_complexity),
                )
            )
        ).scalars().all()
        platform = {row.id: row for row in rows}

    requirements: dict[int, TasDerivedRequirement] = {}
    if requirement_ids:
        rows = (
            await db.execute(
                select(TasDerivedRequirement).where(TasDerivedRequirement.id.in_(requirement_ids))
            )
        ).scalars().all()
        requirements = {row.id: row for row in rows}

    return _Context(sources, platform, requirements, await _load_header_format(db, sources))


async def _load_header_format(
    db: AsyncSession, sources: dict[int, TasSourceTestCase]
) -> dict[str, dict]:
    """Read the header styling back off the workbook these rows came from.

    Read at export rather than stored at import: it is presentation, it is
    always available from the file the platform already keeps, and storing a
    copy would mean a re-uploaded sheet exports with its old colours.

    When several documents fed the export the most common one wins — a mixed
    batch has no single answer, and the majority's styling is the one most of
    the rows were authored under.
    """
    document_ids = [row.source_document_id for row in sources.values() if row.source_document_id]
    if not document_ids:
        return {}

    ranked = sorted(
        {doc_id: document_ids.count(doc_id) for doc_id in set(document_ids)}.items(),
        key=lambda item: (-item[1], item[0]),
    )
    document = await db.get(UploadedDocument, ranked[0][0])
    if document is None or not document.file_path:
        return {}
    try:
        with open(document.file_path, "rb") as handle:
            contents = handle.read()
    except OSError:
        # The styling is a nicety; a missing file must not fail the download.
        return {}
    return sheet_import.read_header_format(contents, document.original_filename or "")


_STUDIO_SOURCE_LABEL = {
    "existing": "Refined from platform test case",
    "imported": "Refined from uploaded test case",
    "derived": "Created to close a coverage gap",
}


def _row(test_case: TasRefinedTestCase, context: _Context, index: int) -> list[str]:
    original = context.source_row(test_case)
    platform = context.platform_case(test_case)
    requirement = context.requirements.get(test_case.derived_requirement_id or -1)

    def inherited(field: str, platform_value: Any = None) -> str:
        """A column the studio does not own: the source's value, or blank.

        Never invented. A gap-derived test case has no Domain or Channel
        because nothing supplied one, and guessing would put a value in the
        team's sheet that no one chose.
        """
        value = original.get(field)
        if value not in (None, ""):
            return _text(value)
        return _text(platform_value)

    template = [
        # `ID` is the sheet's own row number. The uploaded value is kept when
        # there was one so rows line up against the original.
        _text(original.get("id") or index),
        inherited("domain", getattr(platform, "domain_name", None)),
        inherited("channel", getattr(platform, "channel_name", None)),
        inherited("product", getattr(platform, "product_name", None)),
        inherited("area_of_test", getattr(platform, "area_of_test_name", None)),
        # The one column the studio guarantees: the preserved display ID.
        _text(test_case.tc_display_id),
        inherited("environment"),
        inherited("sub_request_type", getattr(platform, "sub_request_type_name", None)),
        # Studio-owned from here: this is the refinement.
        _text(test_case.objective or test_case.title),
        _list_text(test_case.preconditions),
        _steps_text(test_case.steps),
        _text(test_case.expected_result),
        inherited("atc_test_case", getattr(platform, "atc_test_case", None)),
        _text(test_case.test_type) or inherited("test_case_type", getattr(platform, "test_case_type_name", None)),
        inherited("test_case_complexity", getattr(platform, "test_case_complexity_name", None)),
        inherited("tested_by"),
        inherited("jira_or_ppm", getattr(platform, "jira_issue_key", None) or getattr(platform, "ppm_id", None)),
        inherited("overall_status"),
        inherited("blocking_snag"),
        inherited("sit"),
        inherited("planned_execution_sequence"),
        inherited("is_critical", getattr(platform, "is_critical", None)),
        _text(test_case.priority),
        inherited("severity", getattr(platform, "severity", None)),
        _text(test_case.bdd_scenario),
        _text(test_case.status),
        _text(requirement.requirement_key if requirement else original.get("requirement_display_id")),
        _text(requirement.title if requirement else original.get("requirement_title")),
        # The studio has no scenario layer; the columns stay so the shape
        # matches, and stay empty rather than being filled with something else.
        inherited("scenario_display_id"),
        inherited("scenario_title"),
        inherited("test_phase", getattr(platform, "test_phase", None)),
        inherited("product_group", getattr(platform, "product_group", None)),
        inherited("external_tc_id", getattr(platform, "external_tc_id", None)),
        inherited("external_tc_url", getattr(platform, "external_tc_url", None)),
        _text(test_case.created_at.isoformat() if test_case.created_at else None),
    ]

    return template + [
        _text(test_case.classification),
        _text(test_case.classification_reason),
        _text(test_case.application_url),
        _targets_text(test_case.steps),
        _data_keys_text(test_case),
        _text(bool(test_case.test_data_required)),
        _text(test_case.test_data_notes),
        _STUDIO_SOURCE_LABEL.get(test_case.origin, test_case.origin),
        _text(test_case.version),
    ]


def _split(test_cases: list[TasRefinedTestCase], automation: bool) -> list[TasRefinedTestCase]:
    return [
        tc
        for tc in test_cases
        if (tc.classification == "automation") == automation
        # An unclassified test case belongs on the manual sheet: manual is what
        # happens to a test case nobody has automated.
        or (tc.classification not in ("automation", "manual") and not automation)
    ]


def to_csv(test_cases: list[TasRefinedTestCase], context: _Context, *, automation: bool) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(COLUMNS)
    for index, test_case in enumerate(_split(test_cases, automation), start=1):
        writer.writerow(_row(test_case, context, index))
    # BOM so Excel opens the file as UTF-8 rather than mangling non-ASCII.
    return buffer.getvalue().encode("utf-8-sig")


def to_excel(test_cases: list[TasRefinedTestCase], context: _Context) -> bytes:
    """One workbook, one sheet per classification, identical columns.

    Both sheets always exist, even when empty: an automation engineer opening
    the file needs to see that there were no automation test cases, not wonder
    whether the sheet failed to generate.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)

    # Column index (1-based) -> the canonical field it carries, so the uploaded
    # workbook's per-column styling can be replayed onto the same columns.
    field_by_column = {
        index: field for index, (_, field) in enumerate(TEST_CASE_TEMPLATE_COLUMNS, start=1)
    }

    default_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    default_header_fill = PatternFill(fill_type="solid", fgColor="92D050")
    # The studio's own columns are tinted differently so a reader can see at a
    # glance where their sheet ends and this module's additions begin.
    studio_fill = PatternFill(fill_type="solid", fgColor="D99694")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _header_style(column_index: int) -> tuple[Font, PatternFill, float]:
        """The uploaded workbook's styling for this column, or the default.

        Reproducing it matters beyond looks: this template tints each column by
        who owns it, so a reader scans the sheet by colour. A uniformly styled
        download would drop that.
        """
        if column_index > len(TEST_CASE_TEMPLATE_HEADERS):
            return default_header_font, studio_fill, 28
        style = context.header_format.get(field_by_column.get(column_index, ""))
        if not style:
            return default_header_font, default_header_fill, 28
        font = Font(
            name=style.get("name") or "Calibri",
            bold=style.get("bold", True),
            color=style.get("font_color") or "FFFFFF",
            size=style.get("size") or 12,
        )
        fill = (
            PatternFill(fill_type="solid", fgColor=style["fill"])
            if style.get("fill")
            else default_header_fill
        )
        # The source's own width is usually far too narrow for a refined step
        # list, so it sets a floor rather than the value.
        return font, fill, max(float(style.get("width") or 0), 28)
    thin = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="top", wrap_text=True)

    for sheet_name, automation in (("Automation Test Cases", True), ("Manual Test Cases", False)):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(COLUMNS)
        for column_index in range(1, len(COLUMNS) + 1):
            cell = sheet.cell(row=1, column=column_index)
            font, fill, width = _header_style(column_index)
            cell.font = font
            cell.fill = fill
            cell.alignment = header_align
            cell.border = border
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.freeze_panes = "A2"

        for row_index, test_case in enumerate(_split(test_cases, automation), start=2):
            sheet.append(_row(test_case, context, row_index - 1))
            for column_index in range(1, len(COLUMNS) + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = border

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
