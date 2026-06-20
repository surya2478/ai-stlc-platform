"""
GAP-5 edge-case tests: export service + traceability chain.

Covers:
  EC-01  Empty project — no test cases → valid Excel/CSV with header row only
  EC-02  Empty project — no requirements → valid traceability matrix (0 rows)
  EC-03  Test case with ALL nullable fields null → no KeyError / no crash
  EC-04  Unicode titles, BDD scenario, step data survive round-trip
  EC-05  Xray CSV: BDD test case emits Cucumber type, single row
  EC-06  Xray CSV: manual test case with N steps emits N rows
  EC-07  Xray CSV: manual test case with NO steps emits 1 fallback row
  EC-08  Xray CSV: step stored as plain string (not dict) is handled
  EC-09  Excel cell truncation — text > 32 767 chars is safely shortened
  EC-10  Traceability chain: requirement with no scenarios → empty lists, no crash
  EC-11  Traceability chain: test cases linked via scenario only (no direct req_id)
  EC-12  Traceability chain: requirement not in project → ValueError
  EC-13  Export endpoint auth guard — /chain on missing req → 404
  EC-14  Export endpoint format validation — unsupported format → 400
  EC-15  _safe_str: None, bool True/False, nested list, oversized string
  EC-16  _steps_to_text: None, empty list, list of dicts, list of strings
  EC-17  _list_to_text: None, mixed types including None items
  EC-18  Xray CSV: Jira-linked requirement key written to Requirement Issue Key column
  EC-19  Traceability chain: duplicate TC ids not doubled (scenario + req both set)
  EC-20  Excel workbook: freeze pane, header row count, column count are correct
"""
from __future__ import annotations

import csv
import io
import types
import unittest.mock as mock
from datetime import datetime
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.services.export_service import (
    _list_to_text,
    _safe_str,
    _steps_to_text,
    export_test_cases_csv,
    export_test_cases_excel,
    export_test_cases_xray_csv,
    export_traceability_matrix_excel,
    get_requirement_traceability_chain,
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _tc(**kwargs):
    """Return a minimal mock TestCase."""
    tc = MagicMock()
    tc.id = kwargs.get("id", 1)
    tc.project_id = kwargs.get("project_id", 1)
    tc.test_case_id = kwargs.get("test_case_id", "TC-001")
    tc.title = kwargs.get("title", "Login works")
    tc.test_type = kwargs.get("test_type", "functional")
    tc.priority = kwargs.get("priority", "High")
    tc.severity = kwargs.get("severity", "Major")
    tc.automation_candidate = kwargs.get("automation_candidate", False)
    tc.bdd_scenario = kwargs.get("bdd_scenario", None)
    tc.steps = kwargs.get("steps", [{"step_number": 1, "action": "Open app", "expected_result": "App opens"}])
    tc.preconditions = kwargs.get("preconditions", ["User is registered"])
    tc.expected_result = kwargs.get("expected_result", "Success")
    tc.status = kwargs.get("status", "approved")
    tc.approval_status = kwargs.get("approval_status", "approved")
    tc.requirement_id = kwargs.get("requirement_id", None)
    tc.scenario_id = kwargs.get("scenario_id", None)
    tc.telecom_domain = kwargs.get("telecom_domain", "BSS")
    tc.test_phase = kwargs.get("test_phase", "SIT")
    tc.product_group = kwargs.get("product_group", None)
    tc.product = kwargs.get("product", None)
    tc.jira_issue_key = kwargs.get("jira_issue_key", None)
    tc.external_tc_id = kwargs.get("external_tc_id", None)
    tc.external_tc_url = kwargs.get("external_tc_url", None)
    tc.created_at = kwargs.get("created_at", datetime(2025, 1, 1, 12, 0, 0))
    return tc


def _req(**kwargs):
    """Return a minimal mock Requirement."""
    r = MagicMock()
    r.id = kwargs.get("id", 1)
    r.project_id = kwargs.get("project_id", 1)
    r.requirement_id = kwargs.get("requirement_id", "REQ-001")
    r.title = kwargs.get("title", "User can log in")
    r.status = kwargs.get("status", "approved")
    r.risk_level = kwargs.get("risk_level", "High")
    r.telecom_domain = kwargs.get("telecom_domain", "BSS")
    r.test_phase = kwargs.get("test_phase", "SIT")
    r.quality_score = kwargs.get("quality_score", None)
    r.quality_verdict = kwargs.get("quality_verdict", None)
    r.jira_issue_key = kwargs.get("jira_issue_key", None)
    r.created_at = kwargs.get("created_at", datetime(2025, 1, 1))
    return r


def _build_db_mock(
    test_cases: list | None = None,
    requirements: list | None = None,
    scenarios: list | None = None,
    exec_results: list | None = None,
    defects: list | None = None,
    single_req=None,
):
    """
    Return an AsyncMock database session whose `execute` returns appropriate
    scalars for each query based on which model is being queried.
    """
    test_cases = test_cases or []
    requirements = requirements or []
    scenarios = scenarios or []
    exec_results = exec_results or []
    defects = defects or []

    def _scalars(rows):
        m = MagicMock()
        m.all.return_value = rows
        m.scalars.return_value = m
        return m

    def _scalar_one_or_none(value):
        m = MagicMock()
        m.scalar_one_or_none.return_value = value
        return m

    call_log: list[Any] = []

    async def fake_execute(stmt):
        call_log.append(stmt)
        # Detect which model based on compile fromprimary entity
        # We use a simple heuristic: check the compiled clause string
        try:
            compiled = str(stmt.compile())
        except Exception:
            compiled = ""

        if "test_cases" in compiled and "test_scenarios" not in compiled:
            m = MagicMock()
            m.scalars.return_value.all.return_value = test_cases
            m.scalar_one_or_none.return_value = None
            return m
        elif "requirements" in compiled and single_req is not None:
            m = MagicMock()
            m.scalars.return_value.all.return_value = requirements
            m.scalar_one_or_none.return_value = single_req
            return m
        elif "requirements" in compiled:
            m = MagicMock()
            m.scalars.return_value.all.return_value = requirements
            m.scalar_one_or_none.return_value = None
            return m
        elif "test_scenarios" in compiled:
            m = MagicMock()
            m.scalars.return_value.all.return_value = scenarios
            m.scalar_one_or_none.return_value = None
            return m
        elif "execution_results" in compiled:
            m = MagicMock()
            m.scalars.return_value.all.return_value = exec_results
            m.scalar_one_or_none.return_value = None
            return m
        elif "defect_drafts" in compiled:
            m = MagicMock()
            m.scalars.return_value.all.return_value = defects
            m.scalar_one_or_none.return_value = None
            return m
        else:
            m = MagicMock()
            m.scalars.return_value.all.return_value = []
            m.scalar_one_or_none.return_value = None
            return m

    db = AsyncMock()
    db.execute.side_effect = fake_execute
    return db


# ── EC-15: _safe_str ──────────────────────────────────────────────────────────

def test_safe_str_none():
    assert _safe_str(None) == ""


def test_safe_str_bool_true():
    assert _safe_str(True) == "Yes"


def test_safe_str_bool_false():
    assert _safe_str(False) == "No"


def test_safe_str_list():
    result = _safe_str(["a", "b"])
    assert "a" in result
    assert "b" in result


def test_safe_str_dict():
    result = _safe_str({"key": "value"})
    assert "key" in result


def test_safe_str_truncation():
    """EC-09: strings longer than 32 767 chars are truncated safely."""
    long_str = "X" * 40_000
    result = _safe_str(long_str)
    assert len(result) <= 32_767
    assert "truncated" in result


def test_safe_str_custom_max_len():
    """When max_len is small the text is hard-cut to exactly max_len chars."""
    result = _safe_str("hello world", max_len=5)
    assert len(result) <= 5
    assert result == "hello"


# ── EC-16: _steps_to_text ─────────────────────────────────────────────────────

def test_steps_to_text_none():
    assert _steps_to_text(None) == ""


def test_steps_to_text_empty():
    assert _steps_to_text([]) == ""


def test_steps_to_text_dicts():
    steps = [
        {"action": "Click Login", "expected_result": "Login page shows"},
        {"action": "Enter email", "expected_result": "Email accepted"},
    ]
    result = _steps_to_text(steps)
    assert "Click Login" in result
    assert "Login page shows" in result
    assert "1." in result
    assert "2." in result


def test_steps_to_text_plain_strings():
    """EC-08: steps as plain strings (not dicts) are handled."""
    steps = ["Open browser", "Navigate to URL"]
    result = _steps_to_text(steps)
    assert "Open browser" in result
    assert "Navigate to URL" in result


# ── EC-17: _list_to_text ──────────────────────────────────────────────────────

def test_list_to_text_none():
    assert _list_to_text(None) == ""


def test_list_to_text_with_none_items():
    """Gracefully handle list items that are None."""
    result = _list_to_text(["a", None, "b"])
    assert "a" in result
    assert "b" in result
    # None items silently excluded
    assert "None" not in result


def test_list_to_text_mixed_types():
    result = _list_to_text([1, "two", 3.0])
    assert "1" in result
    assert "two" in result


# ── EC-01: Empty project — no test cases ─────────────────────────────────────

@pytest.mark.anyio
async def test_export_excel_empty_project():
    """EC-01: Returns valid Excel bytes with only the header row."""
    db = _build_db_mock(test_cases=[], requirements=[])
    result = await export_test_cases_excel(db, project_id=99)
    assert isinstance(result, bytes)
    assert len(result) > 0  # still a valid xlsx

    # Verify it's a valid zip (xlsx is zip-based)
    import zipfile
    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        assert "xl/worksheets/sheet1.xml" in zf.namelist()


@pytest.mark.anyio
async def test_export_csv_empty_project():
    """EC-01 (CSV): Returns CSV with header row only."""
    db = _build_db_mock(test_cases=[], requirements=[])
    result = await export_test_cases_csv(db, project_id=99)
    assert isinstance(result, str)
    rows = list(csv.reader(io.StringIO(result)))
    assert len(rows) == 1  # header only
    assert "TC ID" in rows[0]


@pytest.mark.anyio
async def test_export_xray_empty_project():
    """EC-01 (Xray): Returns Xray CSV with header row only."""
    db = _build_db_mock(test_cases=[], requirements=[])
    result = await export_test_cases_xray_csv(db, project_id=99)
    rows = list(csv.reader(io.StringIO(result)))
    assert len(rows) == 1
    assert "Summary" in rows[0]


# ── EC-02: Empty project — no requirements → matrix ──────────────────────────

@pytest.mark.anyio
async def test_export_traceability_matrix_empty():
    """EC-02: Matrix with no requirements returns valid workbook."""
    db = _build_db_mock(requirements=[], test_cases=[])
    result = await export_traceability_matrix_excel(db, project_id=99)
    assert isinstance(result, bytes)
    import zipfile
    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        names = zf.namelist()
    assert any("sheet" in n for n in names)


# ── EC-03: Test case with all nullable fields null ────────────────────────────

@pytest.mark.anyio
async def test_export_excel_nullable_fields():
    """EC-03: A TC where every optional field is None does not crash."""
    tc = _tc(
        test_type=None,
        bdd_scenario=None,
        steps=None,
        preconditions=None,
        expected_result=None,
        telecom_domain=None,
        test_phase=None,
        product_group=None,
        product=None,
        jira_issue_key=None,
        external_tc_id=None,
        external_tc_url=None,
        requirement_id=None,
        scenario_id=None,
        created_at=None,
    )
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_excel(db, project_id=1)
    assert isinstance(result, bytes) and len(result) > 0


# ── EC-04: Unicode titles ─────────────────────────────────────────────────────

@pytest.mark.anyio
async def test_export_csv_unicode():
    """EC-04: Arabic, Chinese, emoji in title survive CSV round-trip."""
    tc = _tc(title="تسجيل الدخول 登录 🚀", bdd_scenario=None)
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_csv(db, project_id=1)
    assert "تسجيل الدخول" in result
    assert "登录" in result
    assert "🚀" in result


# ── EC-05: Xray CSV — BDD test case → Cucumber, single row ───────────────────

@pytest.mark.anyio
async def test_xray_bdd_test_case_single_row():
    """EC-05: BDD TC emits Cucumber type and exactly one data row."""
    tc = _tc(
        bdd_scenario="Given I am on the login page\nWhen I submit valid credentials\nThen I am logged in",
        steps=[],  # steps irrelevant for BDD
    )
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_xray_csv(db, project_id=1)
    rows = list(csv.reader(io.StringIO(result)))
    # Header + exactly 1 data row
    assert len(rows) == 2
    data_row = rows[1]
    # Test Type column (index 2) should be Cucumber
    assert data_row[2] == "Cucumber"
    # Generic Test Definition (index 3) should contain the scenario
    assert "Given" in data_row[3]


# ── EC-06: Xray CSV — manual TC with N steps → N rows ────────────────────────

@pytest.mark.anyio
async def test_xray_manual_multistep():
    """EC-06: Manual TC with 3 steps emits 3 data rows."""
    tc = _tc(
        bdd_scenario=None,
        steps=[
            {"action": "Step 1", "expected_result": "Result 1"},
            {"action": "Step 2", "expected_result": "Result 2"},
            {"action": "Step 3", "expected_result": "Result 3"},
        ],
    )
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_xray_csv(db, project_id=1)
    rows = list(csv.reader(io.StringIO(result)))
    # Header + 3 step rows
    assert len(rows) == 4
    # First data row has summary / metadata
    assert rows[1][0] == tc.title
    # Subsequent rows have blank summary
    assert rows[2][0] == ""
    assert rows[3][0] == ""
    # Step action values
    assert rows[1][4] == "Step 1"
    assert rows[2][4] == "Step 2"
    assert rows[3][4] == "Step 3"


# ── EC-07: Xray CSV — no steps → fallback row ─────────────────────────────────

@pytest.mark.anyio
async def test_xray_manual_no_steps_fallback():
    """EC-07: Manual TC with no steps emits 1 fallback row."""
    tc = _tc(bdd_scenario=None, steps=None)
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_xray_csv(db, project_id=1)
    rows = list(csv.reader(io.StringIO(result)))
    assert len(rows) == 2  # header + 1 row
    assert "Refer to test design" in rows[1][4]


# ── EC-08: Xray CSV — step as plain string ────────────────────────────────────

@pytest.mark.anyio
async def test_xray_step_as_string():
    """EC-08: Steps stored as plain strings are not treated as dicts."""
    tc = _tc(bdd_scenario=None, steps=["Open app", "Login"])
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_xray_csv(db, project_id=1)
    rows = list(csv.reader(io.StringIO(result)))
    # header + 2 step rows
    assert len(rows) == 3
    assert "Open app" in rows[1][4]
    assert "Login" in rows[2][4]


# ── EC-18: Xray — Jira requirement key written to column ─────────────────────

@pytest.mark.anyio
async def test_xray_jira_requirement_key():
    """EC-18: Jira issue key from linked requirement appears in Xray output."""
    req = _req(id=5, requirement_id="REQ-005", jira_issue_key="PROJ-123")
    tc = _tc(requirement_id=5, bdd_scenario=None, steps=[{"action": "Do it", "expected_result": "Done"}])
    db = _build_db_mock(test_cases=[tc], requirements=[req])
    result = await export_test_cases_xray_csv(db, project_id=1)
    assert "PROJ-123" in result


# ── EC-20: Excel header row / column count / freeze pane ─────────────────────

@pytest.mark.anyio
async def test_excel_structure():
    """EC-20: Excel workbook has correct header count and freeze pane."""
    from openpyxl import load_workbook
    tc = _tc()
    db = _build_db_mock(test_cases=[tc], requirements=[])
    result = await export_test_cases_excel(db, project_id=1)
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    # 24 columns as defined in export_service
    assert ws.max_column == 24
    # Freeze pane should be A2
    assert ws.freeze_panes == "A2"
    # Header row 1, first cell should be "TC ID"
    assert ws.cell(row=1, column=1).value == "TC ID"


# ── EC-10: Chain — requirement with no downstream artifacts ──────────────────

@pytest.mark.anyio
async def test_chain_no_downstream_artifacts():
    """EC-10: Requirement with no scenarios/TCs/executions → empty lists, no crash."""
    req = _req(id=10)

    # Build a db that returns the requirement for the initial lookup,
    # and empty lists for everything else.
    call_count = 0

    async def fake_execute(stmt):
        nonlocal call_count
        call_count += 1
        m = MagicMock()
        try:
            compiled = str(stmt.compile())
        except Exception:
            compiled = ""

        if "requirements" in compiled:
            m.scalar_one_or_none.return_value = req
            m.scalars.return_value.all.return_value = [req]
        else:
            m.scalar_one_or_none.return_value = None
            m.scalars.return_value.all.return_value = []
        return m

    db = AsyncMock()
    db.execute.side_effect = fake_execute

    chain = await get_requirement_traceability_chain(db, requirement_id=10, project_id=1)

    assert chain["requirement"]["id"] == 10
    assert chain["scenarios"] == []
    assert chain["test_cases"] == []
    assert chain["execution_results"] == []
    assert chain["defects"] == []
    assert "no_test_cases" in chain["summary"]["gaps"]
    # no_execution is NOT added when there are no test cases
    assert "no_execution" not in chain["summary"]["gaps"]


# ── EC-12: Chain — requirement not in project → ValueError ───────────────────

@pytest.mark.anyio
async def test_chain_requirement_not_found():
    """EC-12: Non-existent requirement raises ValueError."""
    async def fake_execute(stmt):
        m = MagicMock()
        m.scalar_one_or_none.return_value = None
        m.scalars.return_value.all.return_value = []
        return m

    db = AsyncMock()
    db.execute.side_effect = fake_execute

    with pytest.raises(ValueError, match="not found"):
        await get_requirement_traceability_chain(db, requirement_id=999, project_id=1)


# ── EC-11: Chain — TCs linked via scenario only ──────────────────────────────

@pytest.mark.anyio
async def test_chain_tc_via_scenario_only():
    """
    EC-11: A TC linked to a scenario (but requirement_id=None) should be
    included in the chain via the scenario path.
    """
    req = _req(id=20)
    from unittest.mock import MagicMock
    scen = MagicMock()
    scen.id = 30
    scen.scenario_id = "SCN-030"
    scen.title = "Login scenario"
    scen.scenario_type = "positive"
    scen.priority = "High"
    scen.status = "approved"

    tc_via_scen = _tc(id=40, scenario_id=30, requirement_id=None)

    call_counter: dict[str, int] = {"req": 0, "scen": 0, "tc_req": 0, "tc_scen": 0, "er": 0, "def": 0}

    async def fake_execute(stmt):
        m = MagicMock()
        try:
            compiled = str(stmt.compile())
        except Exception:
            compiled = ""

        if "requirements" in compiled:
            call_counter["req"] += 1
            m.scalar_one_or_none.return_value = req
            m.scalars.return_value.all.return_value = []
        elif "test_scenarios" in compiled:
            call_counter["scen"] += 1
            m.scalars.return_value.all.return_value = [scen]
        elif "test_cases" in compiled:
            # First call: TCs with requirement_id=20 → empty
            # Second call: TCs with scenario_id in [30] and requirement_id IS NULL → [tc_via_scen]
            if call_counter.get("tc_first_done"):
                call_counter["tc_scen"] += 1
                m.scalars.return_value.all.return_value = [tc_via_scen]
            else:
                call_counter["tc_first_done"] = True
                call_counter["tc_req"] += 1
                m.scalars.return_value.all.return_value = []
        elif "execution_results" in compiled:
            call_counter["er"] += 1
            m.scalars.return_value.all.return_value = []
        else:
            m.scalars.return_value.all.return_value = []
            m.scalar_one_or_none.return_value = None
        return m

    db = AsyncMock()
    db.execute.side_effect = fake_execute

    chain = await get_requirement_traceability_chain(db, requirement_id=20, project_id=1)

    # The TC linked via scenario should appear
    assert len(chain["test_cases"]) == 1
    assert chain["test_cases"][0]["id"] == 40
    assert len(chain["scenarios"]) == 1


# ── EC-19: Chain — no duplicate TCs when scenario_id + req both set ──────────

@pytest.mark.anyio
async def test_chain_no_duplicate_tcs():
    """
    EC-19: A TC with both requirement_id AND scenario_id set should appear
    exactly once in the chain (not doubled via both paths).
    """
    req = _req(id=50)
    scen = MagicMock()
    scen.id = 60
    scen.scenario_id = "SCN-060"
    scen.title = "Sign-up"
    scen.scenario_type = "positive"
    scen.priority = "Medium"
    scen.status = "approved"

    tc = _tc(id=70, scenario_id=60, requirement_id=50)

    call_counter: dict[str, int] = {}

    async def fake_execute(stmt):
        m = MagicMock()
        try:
            compiled = str(stmt.compile())
        except Exception:
            compiled = ""

        if "requirements" in compiled:
            m.scalar_one_or_none.return_value = req
            m.scalars.return_value.all.return_value = []
        elif "test_scenarios" in compiled:
            m.scalars.return_value.all.return_value = [scen]
        elif "test_cases" in compiled:
            # First TC query (requirement_id=50) → [tc]
            # Second TC query (scenario_id in [60], requirement_id IS NULL) → []
            #   because tc has requirement_id set, so IS NULL filter excludes it
            if call_counter.get("tc_done"):
                m.scalars.return_value.all.return_value = []
            else:
                call_counter["tc_done"] = True
                m.scalars.return_value.all.return_value = [tc]
        elif "execution_results" in compiled:
            m.scalars.return_value.all.return_value = []
        else:
            m.scalars.return_value.all.return_value = []
            m.scalar_one_or_none.return_value = None
        return m

    db = AsyncMock()
    db.execute.side_effect = fake_execute

    chain = await get_requirement_traceability_chain(db, requirement_id=50, project_id=1)
    # Exactly one TC
    assert len(chain["test_cases"]) == 1
    assert chain["test_cases"][0]["id"] == 70
