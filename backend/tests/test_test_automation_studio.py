"""Focused tests for the Test Automation Studio module.

DB-free by design, following test_automation_intelligence_edge.py: everything
asserted here is a pure decision — role and navigation visibility, the
manual-only classification rules, the test data contract, the agent's
invariant re-application, and the export shape. The DB-backed flows are
covered by the live walkthrough rather than duplicated with fixtures.

Usage:
    python -m pytest tests/test_test_automation_studio.py -q
    python tests/test_test_automation_studio.py
"""
from __future__ import annotations

import json
import os
import sys
import time
import traceback
from datetime import datetime, timezone

_HERE = os.path.dirname(os.path.abspath(__file__))
_BACKEND_ROOT = os.path.dirname(_HERE)
if _BACKEND_ROOT not in sys.path:
    sys.path.insert(0, _BACKEND_ROOT)

_PASS: list[str] = []
_FAIL: list[tuple[str, str]] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    if cond:
        _PASS.append(name)
    else:
        _FAIL.append((name, detail))


def _run(fn):
    try:
        fn()
    except AssertionError as exc:
        _FAIL.append((fn.__name__, f"AssertionError: {exc}"))
    except Exception as exc:  # noqa: BLE001
        _FAIL.append((fn.__name__, f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"))


from app.models.test_automation_studio import TasRefinedTestCase
from app.models.user import User
from app.services import rbac_service
from app.services.test_automation_studio import classification, export_service, test_data_bridge


def _user(role: str, superuser: bool = False) -> User:
    return User(
        email=f"{role}@example.test",
        full_name=role,
        hashed_password="x",
        role=role,
        is_active=True,
        is_superuser=superuser,
    )


def _test_case(**overrides) -> TasRefinedTestCase:
    defaults = dict(
        project_id=1,
        origin="derived",
        tc_display_id="TC-0001",
        title="Verify login succeeds with valid credentials",
        objective="Confirm a valid user reaches the dashboard.",
        preconditions=["The user account exists and is active."],
        steps=[{"step_number": 1, "action": "Navigate to the login page", "expected_result": "Login form shown"}],
        expected_result="The dashboard is displayed.",
        classification="undecided",
        manual_only_reasons=[],
        test_data_required=False,
        test_data_status="not_required",
        test_data_requirements=[],
        test_data_ids=[],
        status="draft",
        version=1,
        is_current=True,
        edited_by_user=False,
        priority="Medium",
        created_by=1,
        metadata_={},
    )
    defaults.update(overrides)
    return TasRefinedTestCase(**defaults)


# ─── Role and navigation ─────────────────────────────────────────────────────

def test_role_is_registered_globally_and_per_project():
    check(
        "tas.role_in_project_roles",
        rbac_service.TEST_AUTOMATION_USER_ROLE in rbac_service.ROLE_PERMISSIONS,
        "Test_Automation_Users missing from ROLE_PERMISSIONS - the project role picker reads this",
    )
    from app.api.v1.endpoints.users import GLOBAL_ROLES

    check(
        "tas.role_in_global_roles",
        rbac_service.TEST_AUTOMATION_USER_ROLE in GLOBAL_ROLES,
        "Test_Automation_Users missing from GLOBAL_ROLES - it could not be assigned to a user",
    )


def test_role_holds_every_tas_permission():
    perms = rbac_service.ROLE_PERMISSIONS[rbac_service.TEST_AUTOMATION_USER_ROLE]
    for key in rbac_service._GRANULAR_TAS_PERMISSIONS:
        check(f"tas.role_has_{key}", key in perms, f"Test_Automation_Users lacks {key}")


def test_studio_owns_its_upload_route():
    """Regression: Screen 1's upload must not depend on MANAGE_PROJECT.

    The studio originally called POST /documents/upload, which is gated on
    MANAGE_PROJECT — a permission Test_Automation_Users deliberately lacks. The
    result was that the only role the studio exists for could not upload a
    document into it ("Access denied"). The studio now owns an upload route
    gated on tas.intake, and this asserts the two facts that together prevent
    the regression returning.
    """
    from app.api.v1.endpoints import documents, test_automation_studio

    check(
        "tas.platform_upload_still_needs_manage_project",
        documents.MANAGE_PROJECT == rbac_service.MANAGE_PROJECT,
        "the platform upload's guard moved - re-check whether the studio can reuse it",
    )
    paths = {route.path for route in test_automation_studio.router.routes}
    check(
        "tas.has_own_upload_route",
        "/batches/{batch_id}/upload" in paths,
        f"studio upload route missing; routes: {sorted(paths)}",
    )
    perms = rbac_service.ROLE_PERMISSIONS[rbac_service.TEST_AUTOMATION_USER_ROLE]
    check(
        "tas.role_can_intake_without_manage_project",
        rbac_service.TAS_INTAKE in perms and rbac_service.MANAGE_PROJECT not in perms,
        "the studio role must be able to upload without being a project admin",
    )


def test_role_does_not_gain_project_admin_powers():
    # The studio role produces automation assets. Granting it project
    # management or execution rights would make it an administrator by
    # accident, which is the failure this asserts against.
    perms = rbac_service.ROLE_PERMISSIONS[rbac_service.TEST_AUTOMATION_USER_ROLE]
    check("tas.role_no_manage_project", rbac_service.MANAGE_PROJECT not in perms)
    check("tas.role_no_execute_tests", rbac_service.EXECUTE_TESTS not in perms)
    check("tas.role_no_emergency_stop", rbac_service.EXECUTION_EMERGENCY_STOP not in perms)


def test_navigation_groups_match_the_requirement():
    studio_user = _user(rbac_service.TEST_AUTOMATION_USER_ROLE)
    groups = rbac_service.navigation_groups_for_user(studio_user)
    check(
        "tas.nav_exact_set",
        groups
        == [
            rbac_service.NAV_GROUP_TEST_AUTOMATION_STUDIO,
            rbac_service.NAV_GROUP_OPERATIONS,
            rbac_service.NAV_GROUP_SETTINGS,
            rbac_service.NAV_GROUP_OTHERS,
        ],
        f"Studio role navigation is {groups}",
    )


def test_other_roles_keep_todays_menu_without_the_studio():
    for role in ("qa_engineer", "qa_lead", "viewer"):
        groups = rbac_service.navigation_groups_for_user(_user(role))
        check(
            f"tas.nav_{role}_no_studio",
            rbac_service.NAV_GROUP_TEST_AUTOMATION_STUDIO not in groups,
            f"{role} can see the studio group",
        )
        check(
            f"tas.nav_{role}_unchanged",
            len(groups) == len(rbac_service.ALL_NAV_GROUPS) - 1,
            f"{role} navigation changed shape: {groups}",
        )


def test_admin_sees_everything_and_can_reach_the_studio():
    admin = _user("admin", superuser=True)
    check(
        "tas.nav_admin_all",
        rbac_service.navigation_groups_for_user(admin) == list(rbac_service.ALL_NAV_GROUPS),
    )
    check("tas.access_admin", rbac_service.can_access_test_automation_studio(admin))
    check(
        "tas.access_studio_role",
        rbac_service.can_access_test_automation_studio(_user(rbac_service.TEST_AUTOMATION_USER_ROLE)),
    )
    check(
        "tas.access_denied_qa",
        not rbac_service.can_access_test_automation_studio(_user("qa_engineer")),
    )


# ─── Background jobs ─────────────────────────────────────────────────────────

def test_heavy_operations_are_queued_not_served_inline():
    """The three LLM-driven operations must return 202, not their results.

    Each runs for minutes. Served synchronously they outlive the read timeout
    of whatever sits between the browser and the API: the work commits, the
    client sees a failure, and the user retries work that already succeeded.
    """
    from app.main import app

    paths = app.openapi()["paths"]
    queued = {
        "/api/v1/lab/test-automation-studio/batches/{batch_id}/assess",
        "/api/v1/lab/test-automation-studio/projects/{project_id}/test-cases/generate",
        "/api/v1/lab/test-automation-studio/projects/{project_id}/scripts/generate",
    }
    for path in queued:
        check(f"tas.job_route_exists:{path}", path in paths, f"{path} missing")
        if path not in paths:
            continue
        responses = paths[path]["post"]["responses"]
        check(
            f"tas.job_returns_202:{path}",
            "202" in responses,
            f"{path} returns {sorted(responses)} - a long job must be queued",
        )
        check(
            f"tas.job_not_200:{path}",
            "200" not in responses,
            f"{path} still has a 200 path - it would be served inline",
        )


def test_studio_owns_its_job_status_route():
    """The screens must be able to watch a job without audit-log rights.

    The platform's /agent-runs/{id} is gated on VIEW_AUDIT_LOGS, which
    Test_Automation_Users does not hold — polling it returned 403 and the
    progress bar never moved. Widening the role to read a project's audit
    history in order to watch one job is the wrong trade, so the studio serves
    its own status behind `tas.view`, restricted to its own runs.
    """
    from app.api.v1.endpoints import test_automation_studio as tas_endpoints
    from app.main import app

    paths = app.openapi()["paths"]
    check(
        "tas.job_status_route",
        "/api/v1/lab/test-automation-studio/jobs/{agent_run_id}" in paths,
        "studio job-status route missing",
    )
    perms = rbac_service.ROLE_PERMISSIONS[rbac_service.TEST_AUTOMATION_USER_ROLE]
    check(
        "tas.role_lacks_audit_logs",
        rbac_service.VIEW_AUDIT_LOGS not in perms,
        "role gained audit-log rights - the dedicated job route is no longer justified",
    )
    check(
        "tas.role_can_view",
        rbac_service.TAS_VIEW in perms,
        "role cannot poll its own jobs",
    )
    # The route must not be a window onto every agent run in the platform.
    # Pinned as an exact set, not a prefix rule: the whole control is that only
    # the studio's own agents are readable, and a `startswith("tas_")` check
    # would silently admit any future agent that happened to be named that way.
    check(
        "tas.job_status_scoped_to_studio",
        tas_endpoints.STUDIO_AGENT_NAMES
        == {
            "tas_coverage_assessment",
            "tas_test_case_refinement",
            "tas_script_generation",
            "tas_application_discovery",
            "tas_dry_run",
        },
        f"unexpected readable agent names: {tas_endpoints.STUDIO_AGENT_NAMES}",
    )


def test_celery_tasks_are_registered_and_included():
    from app.worker.celery_app import celery_app

    import app.worker.tasks.test_automation_studio_tasks  # noqa: F401

    for name in (
        "tas.assess_coverage",
        "tas.generate_test_cases",
        "tas.generate_scripts",
        "tas.discover_application",
        "tas.dry_run_scripts",
    ):
        check(f"tas.task_registered:{name}", name in celery_app.tasks, f"{name} not registered")

    # A task the worker never imports is a job that queues and never runs.
    includes = celery_app.conf.include or []
    check(
        "tas.task_module_included",
        "app.worker.tasks.test_automation_studio_tasks" in includes,
        f"module missing from celery include: {includes}",
    )


def test_services_expose_request_time_validation():
    """Validation must be callable without running the job.

    The endpoint validates before queueing so an impossible request is refused
    at the click, rather than becoming a queued job that fails a minute later
    in a worker the user cannot see.
    """
    from app.services.test_automation_studio import (
        coverage_service,
        refinement_service,
        script_lab_service,
    )

    check("tas.validate_coverage", callable(coverage_service.prepare_assessment))
    check("tas.validate_documents", callable(coverage_service.validate_documents_ready))
    check("tas.validate_refinement", callable(refinement_service.validate_generation_request))
    check("tas.validate_scripts", callable(script_lab_service.validate_generation_request))
    # And the executors must accept an existing run plus a progress callback,
    # or the run the endpoint handed the client is not the one reporting.
    import inspect

    for fn, name in (
        (coverage_service.execute_assessment, "execute_assessment"),
        (refinement_service.generate_refined_test_cases, "generate_refined_test_cases"),
        (script_lab_service.generate_scripts, "generate_scripts"),
    ):
        params = inspect.signature(fn).parameters
        check(f"tas.{name}_takes_run", "run" in params, f"{name} cannot adopt the endpoint's run")
        check(
            f"tas.{name}_takes_progress",
            "on_progress" in params,
            f"{name} cannot report progress",
        )


def test_documents_ready_validation_rejects_before_queueing():
    from fastapi import HTTPException

    from app.services.test_automation_studio import coverage_service

    try:
        coverage_service.validate_documents_ready([], [])
        check("tas.validate_no_brd", False, "empty batch was accepted")
    except HTTPException as exc:
        check("tas.validate_no_brd", exc.status_code == 422, f"got {exc.status_code}")

    try:
        coverage_service.validate_documents_ready(
            [{"document_id": 1, "filename": "brd.docx", "text": ""}], []
        )
        check("tas.validate_no_text", False, "unextracted document was accepted")
    except HTTPException as exc:
        check("tas.validate_no_text", exc.status_code == 409, f"got {exc.status_code}")

    # A ready document passes.
    coverage_service.validate_documents_ready(
        [{"document_id": 1, "filename": "brd.docx", "text": "A requirement."}], []
    )
    check("tas.validate_ready_ok", True)


# ─── Display ID allocation ───────────────────────────────────────────────────

def test_display_id_sequence_reads_the_highest_in_use():
    from app.services.test_automation_studio.refinement_service import (
        format_display_id,
        highest_display_id_sequence,
    )

    check("tas.seq_empty", highest_display_id_sequence([]) == 0)
    check("tas.seq_none_entries", highest_display_id_sequence([None, "", "junk"]) == 0)
    check(
        "tas.seq_picks_max",
        highest_display_id_sequence(["TC-0001", "TC-0015", "TC-0007"]) == 15,
    )
    # Foreign ID schemes must not be read as this project's sequence, or the
    # next allocation jumps to a number nobody expects.
    check("tas.seq_ignores_other_prefixes", highest_display_id_sequence(["LOGIN_003", "REQ-99"]) == 0)
    check("tas.seq_ignores_suffixed", highest_display_id_sequence(["TC-0004-A"]) == 0)
    check("tas.format_pads", format_display_id(16) == "TC-0016")
    check("tas.format_wide", format_display_id(12345) == "TC-12345")


def test_allocating_several_ids_in_one_run_yields_distinct_values():
    """Regression: every derived test case in a batch got the same ID.

    The allocator used to re-query the max per row, but the session runs with
    autoflush=False, so rows added earlier in the same run were invisible and
    every one of them was handed the identical ID. The insert then died on
    uq_tas_refined_tc_version with a 500. The counter is now held in memory.
    """
    from app.services.test_automation_studio.refinement_service import (
        format_display_id,
        highest_display_id_sequence,
    )

    existing = ["TC-0014", "TC-0015"]
    start = highest_display_id_sequence(existing) + 1
    allocated = [format_display_id(start + offset) for offset in range(3)]
    check("tas.alloc_distinct", len(set(allocated)) == 3, str(allocated))
    check("tas.alloc_continues_sequence", allocated == ["TC-0016", "TC-0017", "TC-0018"], str(allocated))
    check(
        "tas.alloc_no_reuse",
        not set(allocated) & set(existing),
        "allocated an ID already in use",
    )


def test_session_is_not_autoflush_so_the_allocator_cannot_requery():
    """Pins the assumption the regression above was caused by."""
    from app.database import AsyncSessionLocal

    check(
        "tas.session_autoflush_off",
        AsyncSessionLocal.kw.get("autoflush") is False,
        "autoflush changed - re-check whether per-row ID re-querying would now work",
    )


# ─── Classification ──────────────────────────────────────────────────────────

def test_manual_only_condition_forces_manual():
    otp_case = _test_case(
        steps=[
            {
                "step_number": 1,
                "action": "Enter the OTP received by SMS",
                "expected_result": "The OTP is accepted",
            }
        ]
    )
    decision, reason, findings = classification.classify(otp_case, None)
    check("tas.classify_otp_manual", decision == "manual", f"got {decision}")
    check("tas.classify_otp_finding", any(f["code"] == "otp" for f in findings), str(findings))
    check("tas.classify_otp_reason", "manual-only" in reason.lower(), reason)


def test_keyword_matching_respects_word_boundaries():
    # "otp" appears inside "adopted". A substring match would misclassify
    # every test case that happens to contain those three letters.
    case = _test_case(
        steps=[{"step_number": 1, "action": "Verify the adopted tariff is displayed", "expected_result": "Shown"}]
    )
    decision, _, findings = classification.classify(case, None)
    check("tas.classify_no_substring_false_positive", decision == "automation", f"got {decision}")
    check("tas.classify_no_spurious_finding", findings == [], str(findings))


def test_missing_test_data_does_not_force_manual():
    # Readiness is not automatability. A test case blocked only on data is
    # still automatable once the data arrives, and marking it manual would be
    # a permanent answer to a temporary problem.
    case = _test_case(test_data_required=True, test_data_status="needs_user_action")
    decision, reason, _ = classification.classify(case, None)
    check("tas.classify_data_still_automation", decision == "automation", f"got {decision}")
    check("tas.classify_data_reason", "test data" in reason.lower(), reason)


def test_agent_reported_blockers_force_manual():
    case = _test_case(metadata_={"automation_blockers": ["Requires scanning a physical SIM card"]})
    decision, reason, _ = classification.classify(case, None)
    check("tas.classify_blocker_manual", decision == "manual", f"got {decision}")
    check("tas.classify_blocker_reason", "blocker" in reason.lower(), reason)


def test_stepless_test_case_is_manual():
    decision, _, _ = classification.classify(_test_case(steps=[]), None)
    check("tas.classify_no_steps_manual", decision == "manual", f"got {decision}")


# ─── Test data contract (requirement 4) ──────────────────────────────────────

def test_summarize_flags_only_what_needs_a_human():
    required, status, notes = test_data_bridge.summarize(
        [{"key": "username", "resolution": "agent_generated"}]
    )
    check("tas.data_agent_not_required", required is False and status == "agent_provided", status)
    check("tas.data_agent_no_notes", notes is None)

    required, status, notes = test_data_bridge.summarize(
        [
            {"key": "username", "resolution": "agent_generated"},
            {"key": "active_subscriber", "resolution": "existing_record"},
        ]
    )
    check("tas.data_mixed_required", required is True, "a mixed set must still flag the human need")
    check("tas.data_mixed_status", status == "needs_user_action", status)
    check("tas.data_mixed_names_key", "active_subscriber" in (notes or ""), str(notes))

    required, status, notes = test_data_bridge.summarize([])
    check("tas.data_empty", required is False and status == "not_required", status)


def test_test_data_is_created_through_the_shared_service():
    """Regression: the bridge built a TestData row by hand and the insert failed.

    `test_data` has a dozen NOT NULL lifecycle columns (masking_status,
    reservation_status, quality_status, ...) whose defaults live in the DB.
    SQLAlchemy sends explicit NULLs for attributes the model leaves unset, so
    those defaults never applied and the first agent-generated record died on
    a NotNullViolation — mid-job, after the LLM work was already paid for.

    Going through test_data_service is also the promise this module makes: a
    studio-created record must be indistinguishable from a user-created one,
    display id, quality check, lineage and all.
    """
    import inspect

    from app.services.test_automation_studio import test_data_bridge

    source = inspect.getsource(test_data_bridge)
    check(
        "tas.bridge_uses_service",
        "test_data_service.create_test_data" in source,
        "the bridge must create records through test_data_service",
    )
    check(
        "tas.bridge_does_not_construct_rows",
        "TestData(" not in source,
        "the bridge constructs a TestData row directly - NOT NULL columns will be missed",
    )


def test_user_required_resolution_blocks():
    required, status, _ = test_data_bridge.summarize(
        [{"key": "prod_credentials", "resolution": "user_required"}]
    )
    check("tas.data_user_required", required is True and status == "needs_user_action", status)


# ─── Uploaded test cases keep their ID and name ──────────────────────────────

def test_generation_accepts_either_source_and_refuses_neither():
    """Screen 2 has two entry points and needs at least one of them.

    Requiring `requirement_ids` made uploaded test cases unrefinable on their
    own, which is how their ID and name came to be lost: the only reachable
    path minted a new ID from the requirement instead.
    """
    from pydantic import ValidationError

    from app.schemas.test_automation_studio import GenerateRefinedTestCasesRequest

    sources_only = GenerateRefinedTestCasesRequest(source_test_case_ids=[1, 2])
    check("tas.gen_sources_only", sources_only.requirement_ids == [], str(sources_only.requirement_ids))

    requirements_only = GenerateRefinedTestCasesRequest(requirement_ids=[3])
    check(
        "tas.gen_requirements_only",
        requirements_only.source_test_case_ids == [],
        str(requirements_only.source_test_case_ids),
    )

    both = GenerateRefinedTestCasesRequest(requirement_ids=[3], source_test_case_ids=[1])
    check("tas.gen_both", both.requirement_ids == [3] and both.source_test_case_ids == [1])

    try:
        GenerateRefinedTestCasesRequest()
        check("tas.gen_requires_a_source", False, "an empty selection was accepted")
    except ValidationError:
        check("tas.gen_requires_a_source", True)


def test_padding_does_not_hide_a_colliding_number():
    """`TC-01` off a sheet and `TC-0001` from the studio are the same number.

    Reading them as unrelated let the allocator mint `TC-0002` alongside an
    existing `TC-02`. The unique constraint is on the string, so both would
    persist happily while reading as duplicates on the grid.
    """
    from app.services.test_automation_studio.refinement_service import (
        highest_display_id_sequence,
    )

    check("tas.seq_unpadded", highest_display_id_sequence(["TC-01"]) == 1)
    check("tas.seq_padded_equivalent", highest_display_id_sequence(["TC-0001"]) == 1)
    check(
        "tas.seq_mixed_padding_takes_max",
        highest_display_id_sequence(["TC-05", "TC-0002"]) == 5,
        "a sheet ID must be able to win the max",
    )
    # The existing guarantees still hold.
    check("tas.seq_still_ignores_foreign", highest_display_id_sequence(["REQ-99"]) == 0)


def test_origin_vocabulary_covers_an_uploaded_test_case():
    """`imported` is a distinct provenance and must survive the round trip."""
    from app.schemas.test_automation_studio import RefinedTestCaseOut

    origins = RefinedTestCaseOut.model_fields["origin"].annotation
    values = set(getattr(origins, "__args__", ()))
    check("tas.origin_has_imported", "imported" in values, str(values))
    check("tas.origin_keeps_existing", {"existing", "derived"} <= values, str(values))

    # The model default must stay `derived`: a row created without an explicit
    # origin came from a requirement, not from a sheet.
    check(
        "tas.origin_column_default",
        TasRefinedTestCase.__table__.c.origin.default.arg == "derived",
    )


def test_uploaded_test_case_is_addressable():
    """The row that gives an uploaded test case an identity.

    Held only as JSONB inside an assessment, an uploaded test case could not be
    referenced by a refined row, so refinement could reach it solely by
    matching its ID against the platform `test_cases` table — which a project
    using this module normally has not populated.
    """
    from app.models.test_automation_studio import TasSourceTestCase

    columns = TasSourceTestCase.__table__.c
    for column in ("tc_display_id", "title", "batch_id", "project_id"):
        check(f"tas.source_tc_has_{column}", column in columns, f"missing {column}")

    # Scoped to the batch: two intakes may carry the same sheet ID, and
    # re-assessing a batch must update its own rows rather than collide.
    unique = {
        tuple(sorted(col.name for col in constraint.columns))
        for constraint in TasSourceTestCase.__table__.constraints
        if constraint.__class__.__name__ == "UniqueConstraint"
    }
    check(
        "tas.source_tc_unique_per_batch",
        ("batch_id", "tc_display_id") in unique,
        str(unique),
    )

    # And a refined row must be able to point back at it, or the link that
    # preserves the ID and name has nowhere to live.
    check(
        "tas.refined_points_at_source",
        "source_uploaded_test_case_id" in TasRefinedTestCase.__table__.c,
    )


def test_uploaded_test_cases_are_upserted_not_replaced():
    """Re-assessment must not orphan work already refined.

    `tas_refined_test_cases.source_uploaded_test_case_id` is ON DELETE SET
    NULL, so deleting and recreating these rows on every assessment would
    quietly sever the link from a refined test case to the sheet row it came
    from — and the ID it inherited would no longer be traceable to anything.
    """
    import inspect

    from app.services.test_automation_studio import coverage_service

    source = inspect.getsource(coverage_service.sync_source_test_cases)
    check(
        "tas.source_tc_no_bulk_delete",
        "delete()" not in source,
        "sync must not delete existing source rows",
    )
    check(
        "tas.source_tc_matches_on_display_id",
        "by_display" in source,
        "sync must reconcile on the display ID",
    )
    # execute_assessment has to actually call it, or the rows never appear.
    assessment_source = inspect.getsource(coverage_service.execute_assessment)
    check(
        "tas.assessment_syncs_sources",
        "sync_source_test_cases" in assessment_source,
        "assessment does not persist the uploaded test cases",
    )


def test_extracting_test_cases_does_not_require_a_requirements_document():
    """A batch holding only a test case sheet must still reach the workbench.

    Coverage is measured against requirements and rightly needs a BRD or SRD.
    Extraction was welded to it, so "refine the test cases we already have" —
    the case this module exists for — could not start without a requirements
    document the team may not have.
    """
    from fastapi import HTTPException

    from app.services.test_automation_studio import coverage_service

    sheet = [{"document_id": 9, "filename": "tests.xlsx", "text": "TC-01 Login works"}]

    # The coverage gate still rejects a batch with no BRD/SRD ...
    try:
        coverage_service.validate_documents_ready([], sheet)
        check("tas.assess_still_needs_brd", False, "coverage ran without a requirements document")
    except HTTPException as exc:
        check("tas.assess_still_needs_brd", exc.status_code == 422, f"got {exc.status_code}")

    # ... while extraction accepts exactly that batch.
    coverage_service.validate_test_case_documents_ready(sheet)
    check("tas.extract_needs_no_brd", True)

    # It does need a test case document, and one whose text has landed.
    try:
        coverage_service.validate_test_case_documents_ready([])
        check("tas.extract_needs_a_sheet", False, "extraction accepted an empty batch")
    except HTTPException as exc:
        check("tas.extract_needs_a_sheet", exc.status_code == 422, f"got {exc.status_code}")

    try:
        coverage_service.validate_test_case_documents_ready(
            [{"document_id": 9, "filename": "tests.xlsx", "text": ""}]
        )
        check("tas.extract_waits_for_text", False, "extraction ran before text was available")
    except HTTPException as exc:
        check("tas.extract_waits_for_text", exc.status_code == 409, f"got {exc.status_code}")


def test_a_long_document_is_read_in_segments_that_keep_their_header():
    """Regression: a 30-row sheet yielded 19 rows, or none at all.

    One call per document asked for more JSON than the 8000-token output
    budget allows. The response came back cut mid-object, so the whole
    document extracted to nothing. Splitting it fixed that but introduced a
    quieter failure: segments after the first had no column headers, so the
    model picked a different column as the test case ID per segment and
    dropped rows it could not interpret — 11 of 30 lost with the run still
    reporting success. Both are worse than a loud error, because the ID is
    what the studio then preserves forever.
    """
    from app.agents.test_automation_studio.coverage_agent import (
        CHUNK_CHARS,
        _chunk,
        _dedupe_extracted,
        _document_header,
    )

    short = "one line\nanother line\n"
    check("tas.chunk_short_untouched", _chunk(short) == [short], str(_chunk(short)))

    rows = "".join(f"TC-{n:02d} some test case text here\n" for n in range(1, 400))
    chunks = _chunk(rows, size=2_000)
    check("tas.chunk_splits_long", len(chunks) > 1, f"{len(chunks)} chunk(s)")
    check("tas.chunk_respects_size", all(len(c) <= 2_000 for c in chunks))
    # Nothing may be lost or duplicated in the split itself.
    check("tas.chunk_lossless", "".join(chunks) == rows)
    check("tas.chunk_drops_empty_tails", all(c.strip() for c in chunks))
    # And no row may be cut in half, or the model is handed a fragment and
    # invents the missing end of a test case ID.
    check(
        "tas.chunk_keeps_lines_whole",
        all(c.endswith("\n") for c in chunks[:-1]),
        "a segment ended mid-line",
    )

    # A single line longer than the budget still has to go somewhere. The
    # trailing line break lands in a segment of its own and is dropped, so this
    # is five segments of content rather than six with an empty one.
    monster = "x" * 5_000 + "\n"
    monster_chunks = _chunk(monster, size=1_000)
    check("tas.chunk_splits_monster_line", len(monster_chunks) == 5, str(len(monster_chunks)))
    check("tas.chunk_monster_lossless", "".join(monster_chunks) == monster.strip())

    header = _document_header(rows, limit=100)
    check("tas.header_is_the_opening", rows.startswith(header) and header, header[:40])
    check("tas.header_respects_limit", len(header) <= 100, str(len(header)))

    # Repeated header rows across segments must not become duplicate test cases.
    deduped = _dedupe_extracted(
        [
            {"test_case_id": "TC-01", "title": "First"},
            {"test_case_id": "tc-01", "title": "First again"},
            {"test_case_id": None, "title": "By title"},
            {"test_case_id": None, "title": "by title"},
            {"test_case_id": "TC-02", "title": "Second"},
        ]
    )
    check("tas.dedupe_collapses_repeats", len(deduped) == 3, str(deduped))
    check("tas.dedupe_keeps_first", deduped[0]["title"] == "First", str(deduped[0]))

    # The per-call budget has to stay under the output limit for a segment to
    # be parseable at all.
    check("tas.chunk_fits_budget", CHUNK_CHARS <= 12_000, f"CHUNK_CHARS={CHUNK_CHARS}")


def test_one_bad_segment_does_not_lose_the_whole_document():
    """Partial extraction beats none.

    A document is only refused when nothing at all could be read; a segment
    that failed is reported alongside whatever the others returned, so a single
    awkward page costs those rows rather than the entire sheet.
    """
    import inspect

    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent

    source = inspect.getsource(CoverageAssessmentAgent._extract)
    check(
        "tas.extract_returns_partial",
        "if not items:" in source,
        "extraction must only fail when it read nothing",
    )
    # The callers must keep those items rather than skipping the document on
    # any error at all — the bug this pairs with.
    run_source = inspect.getsource(CoverageAssessmentAgent.run)
    extract_source = inspect.getsource(CoverageAssessmentAgent.extract_test_cases)
    for name, body in (("run", run_source), ("extract_test_cases", extract_source)):
        after_error = body.split('doc_errors.append(')[1:]
        check(
            f"tas.{name}_keeps_partial_items",
            all("continue" not in segment.split("\n")[1] for segment in after_error),
            f"{name} discards items when a segment reported an error",
        )


def test_both_entry_points_read_a_template_sheet_the_same_way():
    """Regression: assessing an upload silently lost 3 of 15 test cases.

    The deterministic sheet reader was wired into the standalone extraction but
    not into the coverage assessment, so which button the user pressed decided
    whether their sheet was parsed exactly or paraphrased by a model. Assessing
    went to the model and dropped rows without reporting it.
    """
    import inspect

    from app.services.test_automation_studio import coverage_service

    for name in ("execute_assessment", "execute_test_case_extraction"):
        source = inspect.getsource(getattr(coverage_service, name))
        check(
            f"tas.{name}_parses_template_sheets",
            "parse_template_test_cases" in source,
            f"{name} sends template sheets to the model instead of reading them",
        )

    # Coverage still has to see the rows that never reached the model, or every
    # requirement they cover is reported as a gap.
    assessment = inspect.getsource(coverage_service.execute_assessment)
    check(
        "tas.assessment_feeds_parsed_rows_to_coverage",
        "preparsed_test_cases=" in assessment,
        "parsed test cases are not counted as coverage",
    )


def test_a_wrapped_cell_does_not_become_a_phantom_test_case():
    """Regression: a 15-row sheet parsed to 16 test cases.

    A cell containing line breaks makes the reader emit the continuation as its
    own row — the tail of one expected result arrived as a row reading
    ") remain unaffected.". With no ID it was still kept, and the sync then
    minted an ID for it, so a phantom test case entered the workbench carrying
    a made-up ID the studio would preserve forever.
    """
    import csv
    import io

    from app.services.test_automation_studio import sheet_import

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["ID", "Test Case ID", "Test Case Objective", "Test Steps", "Expected Results"])
    writer.writerow(["1", "TC-01", "Validate the thing", "1. do it", "it works"])
    writer.writerow(["", "", ") remain unaffected.", "", ""])
    writer.writerow(["2", "TC-02", "Validate the other thing", "1. do it", "it works"])
    contents = buffer.getvalue().encode("utf-8")

    rows = sheet_import.parse_sheet(contents, "sheet.csv")
    check("tas.sheet_parsed", rows is not None and len(rows) == 2, str(rows and len(rows)))
    check(
        "tas.sheet_skips_continuation_rows",
        all(row["test_case_id"] for row in rows or []),
        "a row with no ID became a test case",
    )
    check(
        "tas.sheet_keeps_real_ids",
        [row["test_case_id"] for row in rows or []] == ["TC-01", "TC-02"],
        str([row["test_case_id"] for row in rows or []]),
    )


def test_extraction_path_is_queued_like_every_other_llm_pass():
    """It is an LLM call per document, so it cannot be served inline."""
    import inspect

    from app.services.test_automation_studio import coverage_service
    from app.worker.celery_app import celery_app
    import app.worker.tasks.test_automation_studio_tasks  # noqa: F401

    check(
        "tas.extract_task_registered",
        "tas.extract_test_cases" in celery_app.tasks,
        f"registered: {[n for n in celery_app.tasks if n.startswith('tas.')]}",
    )
    params = inspect.signature(coverage_service.execute_test_case_extraction).parameters
    check("tas.extract_takes_run", "run" in params)
    check("tas.extract_takes_progress", "on_progress" in params)

    # And it must not write an assessment: a batch with no requirements has no
    # coverage, and reporting a percentage computed against nothing would be a
    # fabricated number on the screen.
    source = inspect.getsource(coverage_service.execute_test_case_extraction)
    check(
        "tas.extract_records_no_assessment",
        "TasCoverageAssessment(" not in source,
        "extraction fabricated a coverage assessment",
    )
    check(
        "tas.extract_passes_null_assessment",
        "assessment_id=None" in source,
        "extraction must not attribute rows to an assessment that never ran",
    )


def test_refinement_preserves_the_uploaded_id_and_name():
    """The rule this whole change exists to enforce, read off the source.

    Asserted against the branch rather than through a DB round trip, in the
    DB-free spirit of this file: what matters is that the display ID and title
    come from the source row and never from the agent's output.
    """
    import inspect

    from app.services.test_automation_studio import refinement_service

    source = inspect.getsource(refinement_service.generate_refined_test_cases)
    check(
        "tas.refine_uses_uploaded_id",
        "display_id_value = uploaded.tc_display_id" in source,
        "the uploaded display ID is not what gets persisted",
    )
    check(
        "tas.refine_uses_uploaded_title",
        "title = uploaded.title" in source,
        "the uploaded title is not what gets persisted",
    )
    check(
        "tas.refine_marks_imported",
        'origin = "imported"' in source,
        "a refined uploaded test case is not marked as imported",
    )


# ─── Refinement agent invariants (requirement 2b) ────────────────────────────

def test_existing_title_survives_a_paraphrasing_model():
    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    row = agent._normalise(
        {"ref": "r1", "title": "A completely rewritten title", "steps": [], "test_data_requirements": []},
        {"ref": "r1", "mode": "refine", "title": "TC-0007 original name"},
    )
    check("tas.refine_title_locked", row["title"] == "TC-0007 original name", row["title"])


def test_derived_title_is_allowed_to_change():
    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    row = agent._normalise(
        {"ref": "r2", "title": "A better generated title", "steps": [], "test_data_requirements": []},
        {"ref": "r2", "mode": "create", "title": "Draft title"},
    )
    check("tas.refine_derived_title_free", row["title"] == "A better generated title", row["title"])


def test_steps_are_renumbered_contiguously():
    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    row = agent._normalise(
        {
            "ref": "r3",
            "title": "t",
            "steps": [
                {"step_number": 5, "action": "a"},
                {"step_number": 9, "action": "b"},
            ],
            "test_data_requirements": [],
        },
        {"ref": "r3", "mode": "create", "title": "t"},
    )
    check(
        "tas.refine_steps_renumbered",
        [step["step_number"] for step in row["steps"]] == [1, 2],
        str(row["steps"]),
    )


def test_flag_is_recomputed_from_the_declared_requirements():
    # A model that lists an unresolvable data item while setting the flag
    # false would suppress the warning the field exists to raise.
    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    row = agent._normalise(
        {
            "ref": "r4",
            "title": "t",
            "steps": [],
            "test_data_required": False,
            "test_data_requirements": [{"key": "live_account", "resolution": "existing_record"}],
        },
        {"ref": "r4", "mode": "create", "title": "t"},
    )
    check("tas.refine_flag_recomputed", row["test_data_required"] is True)
    check("tas.refine_notes_filled", "live_account" in (row.get("test_data_notes") or ""))


def test_every_studio_llm_call_sets_its_own_output_budget():
    """Regression: a refinement batch came back truncated and was discarded.

    app/llm/provider.py notes that "every agent passes its own" max_tokens —
    an unset value falls back to whatever the route configured, which was low
    enough to cut two test cases off mid-object. The batch then failed JSON
    validation and the run generated nothing while reporting success.

    Asserted by reading the source rather than by mocking a provider: the
    invariant is "no generate() call omits max_tokens", which is a property of
    the call sites, not of any one execution path.
    """
    import ast
    import inspect

    from app.agents.test_automation_studio import coverage_agent, refinement_agent, script_agent

    for module in (coverage_agent, refinement_agent, script_agent):
        short = module.__name__.rsplit(".", 1)[-1]
        tree = ast.parse(inspect.getsource(module))
        calls = [
            node
            for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Attribute)
            and node.func.attr in {"generate", "achat", "acomplete"}
        ]
        check(
            f"tas.{short}_has_calls",
            bool(calls),
            "no LLM calls found - the check would pass vacuously",
        )
        for index, call in enumerate(calls):
            kwargs = {kw.arg for kw in call.keywords}
            check(
                f"tas.{short}_call{index}_sets_max_tokens",
                "max_tokens" in kwargs,
                f"LLM call at line {call.lineno} in {short} omits max_tokens",
            )


def test_every_studio_llm_call_pins_a_temperature():
    """The same document must not read differently on every run.

    None of these agents passed a temperature, so every call sampled at the
    provider's default — 1.0 on an OpenAI-compatible endpoint. Two coverage
    assessments of a byte-identical BRD and test case sheet then split the same
    behaviour into different acceptance criteria and disagreed about whether it
    was covered: one project reported 100% coverage and no gaps, another
    reported 50% and derived a requirement, from the same two files.

    Asserted over the call sites for the same reason as the max_tokens check
    above: the invariant is a property of the calls, not of one execution path.
    """
    import ast
    import inspect

    from app.agents.test_automation_studio import (
        contract_agent,
        coverage_agent,
        refinement_agent,
        script_agent,
    )

    for module in (contract_agent, coverage_agent, refinement_agent, script_agent):
        short = module.__name__.rsplit(".", 1)[-1]
        tree = ast.parse(inspect.getsource(module))
        calls = [
            node
            for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Attribute)
            and node.func.attr in {"generate", "achat", "acomplete"}
        ]
        check(
            f"tas.{short}_has_calls_for_temperature",
            bool(calls),
            "no LLM calls found - the check would pass vacuously",
        )
        for index, call in enumerate(calls):
            kwargs = {kw.arg for kw in call.keywords}
            check(
                f"tas.{short}_call{index}_sets_temperature",
                "temperature" in kwargs,
                f"LLM call at line {call.lineno} in {short} omits temperature",
            )

    # Reading a document is extraction, not invention: the coverage passes take
    # the lowest setting available, and a later edit raising it would restore
    # the spread this fixed.
    check(
        "tas.coverage_temperature_is_lowest",
        coverage_agent.EXTRACTION_TEMPERATURE == 0.0,
        f"got {coverage_agent.EXTRACTION_TEMPERATURE}",
    )


def test_covering_ids_resolve_to_test_cases_that_were_actually_supplied():
    """The matching prompt says copy IDs verbatim; the model does not always.

    Asked about "TC-01_USP Direct_eLife (Fixed)" it answered "TC-01", and
    nothing downstream tolerated that: `covering_test_case_refs` is matched
    against `test_cases.test_case_id` by exact string, so the requirement
    reached Screen 2 with no link to the test case covering it, while the
    coverage screen displayed the shortened ID as though it named something.
    """
    from app.agents.test_automation_studio.coverage_agent import (
        _is_boundary_prefix,
        _resolve_case_ids,
    )

    supplied = [
        {"test_case_id": f"TC-{n:02d}_USP Direct/Indirect_eLife (Fixed)", "title": f"Case {n}"}
        for n in range(1, 16)
    ]

    # The exact shortening observed in assessment 14.
    resolved = _resolve_case_ids(["TC-01", "TC-04", "TC-12"], supplied)
    check(
        "tas.covering_ids_restore_full_id",
        resolved
        == [
            "TC-01_USP Direct/Indirect_eLife (Fixed)",
            "TC-04_USP Direct/Indirect_eLife (Fixed)",
            "TC-12_USP Direct/Indirect_eLife (Fixed)",
        ],
        str(resolved),
    )

    # A verbatim answer — the documented contract — must pass through unchanged.
    verbatim = ["TC-07_USP Direct/Indirect_eLife (Fixed)"]
    check("tas.covering_ids_pass_verbatim", _resolve_case_ids(verbatim, supplied) == verbatim)

    # "TC-1" prefixes TC-10 through TC-15. Guessing one would credit a
    # requirement to a test case nobody assessed, so it resolves to nothing.
    check("tas.covering_ids_reject_ambiguous", _resolve_case_ids(["TC-1"], supplied) == [])

    # An ID naming no supplied test case is worse than an absent one: on screen
    # it reads as evidence.
    check("tas.covering_ids_drop_unknown", _resolve_case_ids(["TC-99", "", "  "], supplied) == [])

    # Case differences and repeats collapse to one reference.
    check(
        "tas.covering_ids_dedupe",
        _resolve_case_ids(["TC-02", "tc-02"], supplied)
        == ["TC-02_USP Direct/Indirect_eLife (Fixed)"],
    )

    # A sheet whose rows carry no ID column is still addressable by title.
    untitled = [{"test_case_id": None, "title": "Verify offer description"}]
    check(
        "tas.covering_ids_fall_back_to_title",
        _resolve_case_ids(["verify offer description"], untitled)
        == ["Verify offer description"],
    )

    check("tas.boundary_prefix_at_separator", _is_boundary_prefix("TC-01", "TC-01_USP"))
    check("tas.boundary_prefix_rejects_mid_token", not _is_boundary_prefix("TC-1", "TC-15"))
    check("tas.boundary_prefix_allows_equal", _is_boundary_prefix("TC-01", "TC-01"))


def test_coverage_state_is_not_claimed_without_evidence():
    """A requirement called "covered" by IDs that resolve to nothing has no
    evidence behind it, and reporting it as covered hides a real gap from the
    derivation pass that exists to close it.

    The downgrade itself is asserted behaviourally in
    `test_coverage_state_follows_the_criteria_the_model_named`. What is checked
    here is that the match pass resolves the IDs at all — the step everything
    downstream depends on, and the one an edit could drop without any single
    row looking wrong.
    """
    import inspect

    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent

    source = inspect.getsource(CoverageAssessmentAgent._assess_coverage)
    check(
        "tas.coverage_resolves_ids",
        "_resolve_case_ids(" in source,
        "the match pass stores the model's IDs without resolving them",
    )


def test_coverage_is_scored_over_acceptance_criteria():
    """The percentage must not be an artifact of how the model split the document.

    Scoring by requirement made it exactly that. The regression case: one BRD
    and one test case sheet, byte-identical across two projects. One assessment
    folded a behaviour into a broader acceptance criterion and reported 100%
    coverage with no gaps; the other stated it as its own criterion, found
    nothing exercising it, and reported a 50% gap — because a single extracted
    requirement can only ever score 0, 50 or 100. Counting criteria, the same
    finding is 80%: four of five exercised.
    """
    from app.services.test_automation_studio.coverage_service import score_coverage

    percent, total, covered = score_coverage(
        [{"total_criteria": 5, "covered_criteria_count": 4, "coverage_state": "partially_covered"}],
        1,
    )
    check(
        "tas.score_counts_criteria",
        (percent, total, covered) == (80, 5, 4),
        str((percent, total, covered)),
    )

    # Across requirements it is the criteria that add up, not the requirements:
    # a requirement with nine criteria weighs more than one with a single
    # criterion, which is the whole point of counting them.
    percent, total, covered = score_coverage(
        [
            {"total_criteria": 9, "covered_criteria_count": 9, "coverage_state": "covered"},
            {"total_criteria": 1, "covered_criteria_count": 0, "coverage_state": "uncovered"},
        ],
        2,
    )
    check("tas.score_weighs_by_criteria", percent == 90, f"got {percent}")

    check(
        "tas.score_all_covered",
        score_coverage([{"total_criteria": 3, "covered_criteria_count": 3}], 1)[0] == 100,
    )
    check(
        "tas.score_none_covered",
        score_coverage([{"total_criteria": 3, "covered_criteria_count": 0}], 1)[0] == 0,
    )

    # No requirement listed a criterion: nothing finer to count, so the old
    # requirement-level share stands rather than reporting a false 0%.
    percent, total, covered = score_coverage(
        [
            {"coverage_state": "covered", "total_criteria": 0},
            {"coverage_state": "partially_covered", "total_criteria": 0},
        ],
        2,
    )
    check(
        "tas.score_falls_back_by_requirement",
        (percent, total, covered) == (75, 0, 0),
        str(percent),
    )
    check("tas.score_empty_is_zero", score_coverage([], 0) == (0, 0, 0))


def test_criterion_numbers_are_validated_against_the_requirement():
    """A criterion number the requirement does not have would inflate the score.

    That is the one direction a coverage figure must never drift: reporting
    more tested than was judged sends untested behaviour to automation as
    though it were covered.
    """
    from app.agents.test_automation_studio.coverage_agent import (
        _resolve_criteria,
        _state_from_criteria,
    )

    check("tas.criteria_keep_valid", _resolve_criteria([1, 3], 5) == [1, 3])
    check("tas.criteria_drop_past_end", _resolve_criteria([1, 9], 5) == [1])
    check("tas.criteria_drop_zero_and_negative", _resolve_criteria([0, -1, 2], 5) == [2])
    check("tas.criteria_dedupe", _resolve_criteria([2, 2, 2], 5) == [2])
    check("tas.criteria_sorted", _resolve_criteria([4, 1, 3], 5) == [1, 3, 4])
    check("tas.criteria_ignore_non_numbers", _resolve_criteria(["x", None, 1], 5) == [1])
    check("tas.criteria_none_when_no_criteria", _resolve_criteria([1, 2], 0) == [])

    # State is a function of the counts, not a separate opinion the model can
    # disagree with itself about.
    check("tas.state_all", _state_from_criteria(5, 5) == "covered")
    check("tas.state_none", _state_from_criteria(0, 5) == "uncovered")
    check("tas.state_some", _state_from_criteria(4, 5) == "partially_covered")
    # More covered than exist can only be a bug upstream; it must still not
    # read as anything other than fully covered.
    check("tas.state_over", _state_from_criteria(7, 5) == "covered")


def test_coverage_state_follows_the_criteria_the_model_named():
    """End to end through the matching pass, because the failure was the two
    disagreeing: the model emitted a coverage_state and a criteria judgement
    independently, so a row could read "covered" while naming a gap."""
    import asyncio

    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent

    requirements = [
        {
            "title": "Update eContract offer description",
            "acceptance_criteria": ["one", "two", "three", "four", "five"],
        }
    ]
    existing = [
        {"test_case_id": f"TC-{n:02d}_USP Direct_eLife (Fixed)", "title": f"case {n}"}
        for n in range(1, 4)
    ]

    class _FakeLLM:
        def __init__(self, payload):
            self.payload = payload

        async def generate(self, system, user, **kwargs):
            return json.dumps(self.payload)

    def _assess(payload):
        agent = CoverageAssessmentAgent.__new__(CoverageAssessmentAgent)
        agent._logs = []
        agent.llm = _FakeLLM(payload)
        return asyncio.run(agent._assess_coverage(requirements, existing))

    # The model claims full coverage while naming only four of five criteria.
    # The criteria win.
    rows = _assess(
        [
            {
                "requirement_title": "Update eContract offer description",
                "coverage_state": "covered",
                "covered_criteria": [1, 2, 3, 4],
                "covering_test_case_ids": ["TC-01", "TC-02"],
            }
        ]
    )
    check(
        "tas.state_derived_not_claimed",
        rows[0]["coverage_state"] == "partially_covered",
        str(rows[0]),
    )
    check(
        "tas.row_reports_totals",
        (rows[0]["covered_criteria_count"], rows[0]["total_criteria"]) == (4, 5),
        str(rows[0]),
    )
    check("tas.row_states_the_gap", bool(rows[0]["gap_reason"]), "a gap row with no reason")
    # And the shortened IDs still resolve to the test cases actually supplied.
    check(
        "tas.row_resolves_ids",
        rows[0]["covering_test_case_ids"]
        == ["TC-01_USP Direct_eLife (Fixed)", "TC-02_USP Direct_eLife (Fixed)"],
        str(rows[0]["covering_test_case_ids"]),
    )

    # Every criterion named, and the row reads covered with no gap reason left
    # over from a previous state.
    rows = _assess(
        [
            {
                "requirement_title": "Update eContract offer description",
                "coverage_state": "partially_covered",
                "covered_criteria": [1, 2, 3, 4, 5],
                "covering_test_case_ids": ["TC-01"],
                "gap_reason": "stale reason",
            }
        ]
    )
    check("tas.state_covered_when_all_named", rows[0]["coverage_state"] == "covered", str(rows[0]))
    check(
        "tas.covered_row_has_no_gap_reason",
        rows[0]["gap_reason"] is None,
        str(rows[0]["gap_reason"]),
    )

    # Criteria credited to test cases that do not exist are not credited at all.
    rows = _assess(
        [
            {
                "requirement_title": "Update eContract offer description",
                "coverage_state": "covered",
                "covered_criteria": [1, 2, 3, 4, 5],
                "covering_test_case_ids": ["TC-99", "SOMETHING-ELSE"],
            }
        ]
    )
    check("tas.no_evidence_no_coverage", rows[0]["coverage_state"] == "uncovered", str(rows[0]))
    check("tas.no_evidence_no_criteria", rows[0]["covered_criteria_count"] == 0, str(rows[0]))


def test_gap_derivation_is_told_which_criteria_are_missing():
    """Sending the whole criteria list made the derivation pass re-derive which
    were missing — work the matching pass had already done — and it proposed
    requirements for behaviour that was already tested."""
    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent

    payload = CoverageAssessmentAgent._gap_payload(
        {
            "requirement_title": "R",
            "coverage_state": "partially_covered",
            "covered_criteria": [1, 3],
        },
        {"r": {"acceptance_criteria": ["one", "two", "three", "four"]}},
    )
    check(
        "tas.gap_lists_uncovered",
        [item["text"] for item in payload["uncovered_acceptance_criteria"]] == ["two", "four"],
        str(payload["uncovered_acceptance_criteria"]),
    )
    check(
        "tas.gap_lists_covered",
        [item["text"] for item in payload["already_covered_acceptance_criteria"]]
        == ["one", "three"],
        str(payload["already_covered_acceptance_criteria"]),
    )
    check(
        "tas.gap_keeps_numbering",
        [item["n"] for item in payload["uncovered_acceptance_criteria"]] == [2, 4],
        str(payload["uncovered_acceptance_criteria"]),
    )


def test_the_model_that_answered_a_run_is_the_one_recorded():
    """`agent_runs.llm_model` recorded what was requested, never what answered.

    OpenRouter reroutes to OPENROUTER_FALLBACK_MODELS when the primary is rate
    limited or errors, so a run could be served by a different model entirely —
    which is one reason two assessments of one document disagreed — and the
    audit trail showed the same model for both.
    """
    import asyncio

    from app.llm.provider import _record_served_model, observe_served_models

    # Outside a block, recording is a no-op rather than an error: most callers
    # are not being observed.
    _record_served_model("nobody-is-listening")

    with observe_served_models() as served:
        _record_served_model("deepseek/deepseek-v4-flash")
        _record_served_model("deepseek/deepseek-v4-flash")
        _record_served_model("anthropic/claude-sonnet-4.6")
        _record_served_model(None)
        _record_served_model("   ")
    check(
        "tas.served_models_collected",
        served == {"deepseek/deepseek-v4-flash", "anthropic/claude-sonnet-4.6"},
        str(served),
    )

    # Nesting records into every enclosing observer. The inner block keeps its
    # own view, and the outer one still sees what answered inside it — an inner
    # block that captured calls for itself alone would leave the run wrapping
    # it reporting no model at all.
    with observe_served_models() as outer:
        _record_served_model("outer-model")
        with observe_served_models() as inner:
            _record_served_model("inner-model")
        _record_served_model("outer-again")
    check("tas.served_models_nest_inner", inner == {"inner-model"}, str(inner))
    check(
        "tas.served_models_nest_outer",
        outer == {"outer-model", "inner-model", "outer-again"},
        str(outer),
    )

    # The collector is per-context, so concurrent runs in one worker process
    # cannot attribute one run's model to another.
    async def _one(name: str) -> set[str]:
        with observe_served_models() as observed:
            await asyncio.sleep(0)
            _record_served_model(name)
            await asyncio.sleep(0)
        return observed

    async def _both():
        return await asyncio.gather(_one("model-a"), _one("model-b"))

    first, second = asyncio.run(_both())
    check(
        "tas.served_models_isolated_per_task",
        (first, second) == ({"model-a"}, {"model-b"}),
        str((first, second)),
    )


def test_every_studio_task_records_what_served_it():
    """The observation is worth nothing if the tasks do not enter it. Asserted
    over the module because the invariant is "every task", and a seventh added
    later without it would report a model it never used."""
    import ast
    import inspect

    from app.worker.tasks import test_automation_studio_tasks as tasks

    tree = ast.parse(inspect.getsource(tasks))
    entered = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.AsyncFunctionDef)
        and any(
            isinstance(item, ast.Call)
            and isinstance(item.func, ast.Name)
            and item.func.id == "_agent_run"
            for item in ast.walk(node)
        )
    ]
    # Every coroutine that owns a run takes the session from _agent_run, which
    # is what enters the observation and writes the result back.
    owners = [
        node.name
        for node in ast.walk(tree)
        if isinstance(node, ast.AsyncFunctionDef) and "agent_run_id" in
        [arg.arg for arg in node.args.args]
        and node.name not in {"_load_run", "_fail", "_agent_run"}
    ]
    check("tas.tasks_found", len(owners) >= 6, f"found {owners}")
    check(
        "tas.every_task_observes_its_models",
        {node.name for node in entered} >= set(owners),
        f"tasks not observing served models: {set(owners) - {node.name for node in entered}}",
    )


def test_unchanged_documents_are_not_read_again():
    """Re-extracting on every assessment is what kept the score moving.

    Pinning the temperature narrowed the spread but could not close it: a
    hosted model is not deterministic even at 0, and three consecutive runs
    over one byte-identical BRD split it into 5, 8 and 6 acceptance criteria.
    A percentage cannot settle while its own denominator is re-sampled. With
    the requirements carried over, five consecutive runs of the real batch
    returned 80% and the same four criteria every time.
    """
    import asyncio

    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent
    from app.services.test_automation_studio.coverage_service import requirement_fingerprint

    # The fingerprint is what decides. Same words, same fingerprint; a changed
    # document, a new one — otherwise an edited BRD would never be re-read.
    docs = [{"document_id": 1, "text": "the requirement"}, {"document_id": 2, "text": "another"}]
    check(
        "tas.fingerprint_is_stable",
        requirement_fingerprint(docs) == requirement_fingerprint(docs),
    )
    check(
        "tas.fingerprint_ignores_order",
        requirement_fingerprint(docs) == requirement_fingerprint(list(reversed(docs))),
        "the same documents in a different order must not read as a change",
    )
    check(
        "tas.fingerprint_follows_the_text",
        requirement_fingerprint(docs)
        != requirement_fingerprint(
            [{"document_id": 1, "text": "the requirement, amended"}, docs[1]]
        ),
        "an edited document must be read again",
    )
    check(
        "tas.fingerprint_follows_the_document_set",
        requirement_fingerprint(docs) != requirement_fingerprint(docs[:1]),
        "a removed document must be read again",
    )

    # Given requirements, the extraction pass must not run at all: re-reading
    # would replace the stable set with a freshly sampled one.
    extraction_calls = 0

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            nonlocal extraction_calls
            if "business analyst" in system:
                extraction_calls += 1
                return json.dumps([{"title": "freshly sampled", "acceptance_criteria": ["x"]}])
            if "assessing whether a set of existing test cases" in system:
                return json.dumps(
                    [
                        {
                            "requirement_title": "carried over",
                            "covered_criteria": [1],
                            "covering_test_case_ids": ["TC-01"],
                        }
                    ]
                )
            return json.dumps([])

    agent = CoverageAssessmentAgent.__new__(CoverageAssessmentAgent)
    agent._logs = []
    agent.llm = _FakeLLM()
    agent._call_budget = None

    result = asyncio.run(
        agent.run(
            requirement_documents=[{"document_id": 1, "filename": "brd.docx", "text": "words"}],
            test_case_documents=[],
            preparsed_test_cases=[{"test_case_id": "TC-01", "title": "a case"}],
            preextracted_requirements=[
                {"title": "carried over", "acceptance_criteria": ["one", "two"]}
            ],
            derive_gap_requirements=False,
        )
    )

    check("tas.carryover_succeeds", result.success is True, str(result.error))
    check(
        "tas.carryover_skips_extraction",
        extraction_calls == 0,
        f"the BRD was read again ({extraction_calls} extraction call(s))",
    )
    check(
        "tas.carryover_keeps_the_requirements",
        [req["title"] for req in result.data["requirements"]] == ["carried over"],
        str(result.data["requirements"]),
    )
    # And the carried-over criteria are what coverage is then scored against.
    row = result.data["coverage_rows"][0]
    check(
        "tas.carryover_scores_against_them",
        (row["covered_criteria_count"], row["total_criteria"]) == (1, 2),
        str(row),
    )

    # Without them the document is read as before — the fallback that keeps a
    # first assessment, or one whose documents changed, working.
    agent = CoverageAssessmentAgent.__new__(CoverageAssessmentAgent)
    agent._logs = []
    agent.llm = _FakeLLM()
    agent._call_budget = None
    asyncio.run(
        agent.run(
            requirement_documents=[{"document_id": 1, "filename": "brd.docx", "text": "words"}],
            test_case_documents=[],
            preparsed_test_cases=[{"test_case_id": "TC-01", "title": "a case"}],
            derive_gap_requirements=False,
        )
    )
    check(
        "tas.no_carryover_still_extracts",
        extraction_calls == 1,
        f"expected one extraction call, got {extraction_calls}",
    )


def test_coverage_is_the_majority_judgement_not_one_sample():
    """Judging coverage is an opinion, and a hosted model does not hold one still.

    Measured on the real batch, with the requirement set already fixed so the
    denominator could not move: asked eight times about one requirement of five
    acceptance criteria, the model answered "four covered" five times, "two"
    twice and "all five" once — 40%, 80% and 100% from identical input. Neither
    the pinned temperature nor the carried-over requirements close that, because
    the variance is in the judgement itself. Taking each criterion by majority
    does: the three samples below are the three answers actually observed.
    """
    from app.agents.test_automation_studio.coverage_agent import _vote

    observed = [
        {
            "requirement_title": "Update Offer Description Source in eContract from ECM",
            "covered_criteria": [1, 2, 3, 4],
            "covering_test_case_ids": ["TC-01", "TC-02"],
            "coverage_state": "partially_covered",
            "gap_reason": "criterion 5 is not exercised",
        },
        {
            "requirement_title": "Update Offer Description Source in eContract from ECM",
            "covered_criteria": [3, 4],
            "covering_test_case_ids": ["TC-01"],
            "coverage_state": "partially_covered",
            "gap_reason": "criteria 1, 2 and 5 are not exercised",
        },
        {
            "requirement_title": "Update Offer Description Source in eContract from ECM",
            "covered_criteria": [1, 2, 3, 4, 5],
            "covering_test_case_ids": ["TC-01", "TC-02", "TC-09"],
            "coverage_state": "covered",
            "gap_reason": None,
        },
    ]
    row = _vote([[sample] for sample in observed])[0]

    check(
        "tas.vote_takes_the_majority",
        row["covered_criteria"] == [1, 2, 3, 4],
        str(row["covered_criteria"]),
    )
    # An ID only one sample of three named is not evidence, it is noise.
    check(
        "tas.vote_drops_minority_ids",
        row["covering_test_case_ids"] == ["TC-01", "TC-02"],
        str(row["covering_test_case_ids"]),
    )
    check("tas.vote_takes_state_by_majority", row["coverage_state"] == "partially_covered", str(row))
    # The sentence on the row has to describe the row, not a sample that lost.
    check(
        "tas.vote_reason_matches_the_verdict",
        row["gap_reason"] == "criterion 5 is not exercised",
        str(row["gap_reason"]),
    )

    # Unanimity is unchanged, and a single sample is passed through untouched
    # so setting the sample count to 1 costs nothing but the extra calls.
    unanimous = [[{"requirement_title": "R", "covered_criteria": [1, 2]}]] * 3
    check("tas.vote_unanimous", _vote(unanimous)[0]["covered_criteria"] == [1, 2])
    single = [{"requirement_title": "R", "covered_criteria": [1], "coverage_state": "x"}]
    check("tas.vote_single_sample_untouched", _vote([single]) == single)

    # A minority never wins. With an odd number of samples a tie cannot arise
    # at all, which is why the vote is always taken over one — see
    # `test_an_even_number_of_samples_still_has_a_majority`.
    minority = [
        [{"requirement_title": "R", "covered_criteria": [1, 2]}],
        [{"requirement_title": "R", "covered_criteria": [1]}],
        [{"requirement_title": "R", "covered_criteria": [1]}],
    ]
    check(
        "tas.vote_minority_loses",
        _vote(minority)[0]["covered_criteria"] == [1],
        str(_vote(minority)[0]["covered_criteria"]),
    )

    # A sample that omits a requirement must not shift the others: rows are
    # keyed by title, and the missing vote simply is not cast.
    partial = [
        [
            {"requirement_title": "A", "covered_criteria": [1]},
            {"requirement_title": "B", "covered_criteria": [1]},
        ],
        [{"requirement_title": "A", "covered_criteria": [1]}],
        [
            {"requirement_title": "A", "covered_criteria": [1]},
            {"requirement_title": "B", "covered_criteria": [2]},
        ],
    ]
    merged = {row["requirement_title"]: row["covered_criteria"] for row in _vote(partial)}
    check("tas.vote_keeps_rows_independent", merged == {"A": [1], "B": []}, str(merged))
    check("tas.vote_keeps_every_requirement", len(_vote(partial)) == 2, str(_vote(partial)))

    # Titles differing only in case are one requirement, matching how the rest
    # of this module decides identity.
    cased = [
        [{"requirement_title": "Same Thing", "covered_criteria": [1]}],
        [{"requirement_title": "same thing", "covered_criteria": [1]}],
        [{"requirement_title": "SAME THING", "covered_criteria": [1]}],
    ]
    check("tas.vote_matches_titles_case_insensitively", len(_vote(cased)) == 1, str(_vote(cased)))


def test_the_merged_answer_has_one_order():
    """The merge counted votes through a set, so the order of the covering IDs
    followed string hashing and differed between processes — in the one
    function whose entire purpose is an answer that repeats. Two IDs on equal
    votes must come back in the order they were proposed, every time.
    """
    from app.agents.test_automation_studio.coverage_agent import _vote

    samples = [
        [
            {
                "requirement_title": "R",
                "covered_criteria": [1],
                "covering_test_case_ids": ["TC-01", "TC-02", "TC-03"],
            }
        ]
    ] * 3
    first = _vote(samples)[0]["covering_test_case_ids"]
    check(
        "tas.vote_order_follows_proposal",
        first == ["TC-01", "TC-02", "TC-03"],
        str(first),
    )
    check(
        "tas.vote_order_is_repeatable",
        all(_vote(samples)[0]["covering_test_case_ids"] == first for _ in range(20)),
        "the merged order changed between identical calls",
    )
    # More votes still outrank an earlier proposal.
    ranked = _vote(
        [
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-09", "TC-01"]}],
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-01"]}],
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-01"]}],
        ]
    )[0]["covering_test_case_ids"]
    check("tas.vote_order_prefers_more_votes", ranked == ["TC-01"], str(ranked))

    # A sample repeating one ID must not thereby out-vote the others.
    repeated = _vote(
        [
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-07", "TC-07", "TC-07"]}],
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-08"]}],
            [{"requirement_title": "R", "covering_test_case_ids": ["TC-08"]}],
        ]
    )[0]["covering_test_case_ids"]
    check("tas.vote_counts_one_per_sample", repeated == ["TC-08"], str(repeated))


def test_an_even_number_of_samples_still_has_a_majority():
    """Observed live: of three samples one was truncated and discarded, and the
    two survivors — [1,2,3,4] and [] — merged to [], reporting 0% coverage that
    neither sample had reported. "More than half of two" is unanimity, not a
    majority. Voting over an odd number keeps the answer to something a sample
    actually said.
    """
    from app.agents.test_automation_studio.coverage_agent import _vote

    two = [
        [{"requirement_title": "R", "covered_criteria": [1, 2, 3, 4]}],
        [{"requirement_title": "R", "covered_criteria": []}],
    ]
    check(
        "tas.vote_even_is_not_intersection",
        _vote(two)[0]["covered_criteria"] == [1, 2, 3, 4],
        str(_vote(two)[0]["covered_criteria"]),
    )

    # Four samples vote as three, so the result stays a real majority rather
    # than a tie broken by nothing.
    four = [
        [{"requirement_title": "R", "covered_criteria": [1]}],
        [{"requirement_title": "R", "covered_criteria": [1]}],
        [{"requirement_title": "R", "covered_criteria": [2]}],
        [{"requirement_title": "R", "covered_criteria": [2]}],
    ]
    check("tas.vote_even_reduces_to_odd", _vote(four)[0]["covered_criteria"] == [1], str(_vote(four)))


def test_coverage_samples_survive_one_bad_call():
    """Two agreeing answers still decide. Failing the pass because one of three
    samples timed out would trade the variance this fixes for an outage."""
    import asyncio

    from app.agents.test_automation_studio.coverage_agent import CoverageAssessmentAgent

    calls = 0

    class _FlakyLLM:
        async def generate(self, system, user, **kwargs):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise RuntimeError("upstream hiccup")
            return json.dumps(
                [
                    {
                        "requirement_title": "R",
                        "covered_criteria": [1, 2],
                        "covering_test_case_ids": ["TC-01"],
                    }
                ]
            )

    agent = CoverageAssessmentAgent.__new__(CoverageAssessmentAgent)
    agent._logs = []
    agent.llm = _FlakyLLM()
    agent._call_budget = None

    rows = asyncio.run(
        agent._assess_coverage(
            [{"title": "R", "acceptance_criteria": ["one", "two", "three"]}],
            [{"test_case_id": "TC-01", "title": "a case"}],
        )
    )
    check("tas.samples_run_more_than_once", calls > 1, f"only {calls} call(s) - not sampled")
    check(
        "tas.samples_tolerate_one_failure",
        rows[0]["covered_criteria_count"] == 2,
        str(rows[0]),
    )
    check(
        "tas.samples_survive_to_a_real_state",
        rows[0]["coverage_state"] == "partially_covered",
        str(rows[0]),
    )

    # All samples failing is still the documented fallback, not a crash.
    class _DeadLLM:
        async def generate(self, system, user, **kwargs):
            raise RuntimeError("provider down")

    agent = CoverageAssessmentAgent.__new__(CoverageAssessmentAgent)
    agent._logs = []
    agent.llm = _DeadLLM()
    agent._call_budget = None
    rows = asyncio.run(
        agent._assess_coverage(
            [{"title": "R", "acceptance_criteria": ["one"]}],
            [{"test_case_id": "TC-01", "title": "a case"}],
        )
    )
    check("tas.all_samples_failing_falls_back", rows[0]["assessment_failed"] is True, str(rows[0]))
    check("tas.all_samples_failing_is_uncovered", rows[0]["coverage_state"] == "uncovered", str(rows[0]))


def test_refinement_batch_size_stays_modest():
    """A batch is discarded whole when its response is truncated, so the batch
    size is a blast radius as much as a throughput knob."""
    from app.agents.test_automation_studio.refinement_agent import (
        REFINE_BATCH_SIZE,
        REFINE_MAX_TOKENS,
    )

    check("tas.refine_batch_bounded", 1 <= REFINE_BATCH_SIZE <= 3, f"got {REFINE_BATCH_SIZE}")
    check("tas.refine_budget", REFINE_MAX_TOKENS >= 4000, f"got {REFINE_MAX_TOKENS}")


def test_refinement_calls_overlap_and_still_return_in_input_order():
    """Refining strictly in order left the run waiting on one round trip at a
    time — fifteen test cases took nine and a half minutes.

    Two properties, because the fix trades one for the other if done carelessly:
    the calls must actually overlap, and the results must still come back in
    input order regardless of which call finishes first. The fake provider
    below makes the first item the slowest, so an implementation that appended
    results as they arrived would return them reordered.
    """
    import asyncio

    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    in_flight = 0
    peak_in_flight = 0

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            nonlocal in_flight, peak_in_flight
            in_flight += 1
            peak_in_flight = max(peak_in_flight, in_flight)
            try:
                payload = json.loads(user.split("<user_content>")[1].split("</user_content>")[0])
                ref = payload["items"][0]["ref"]
                # Earlier items answer slower, so completion order is reversed.
                await asyncio.sleep(0.05 * (5 - int(ref[1:])))
                return json.dumps([{"ref": ref, "title": f"title {ref}", "steps": []}])
            finally:
                in_flight -= 1

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    agent._logs = []
    agent.llm = _FakeLLM()

    items = [{"ref": f"r{n}", "mode": "create", "title": f"title {n}"} for n in range(1, 5)]
    reports: list[tuple[int, int]] = []

    async def _on_item(done: int, total: int, label: str) -> None:
        reports.append((done, total))

    result = asyncio.run(agent.run(items=items, on_item=_on_item, concurrency=4))

    check("tas.refine_concurrent_success", result.success is True, str(result.error))
    check(
        "tas.refine_calls_overlap",
        peak_in_flight > 1,
        f"peak in-flight was {peak_in_flight} - the calls ran one at a time",
    )
    check(
        "tas.refine_order_preserved",
        [row["ref"] for row in result.data["refined"]] == ["r1", "r2", "r3", "r4"],
        str([row["ref"] for row in result.data["refined"]]),
    )
    check(
        "tas.refine_progress_counts_completions",
        [done for done, _ in reports] == [1, 2, 3, 4],
        str(reports),
    )


def test_refinement_concurrency_is_bounded_by_configuration():
    """Unbounded fan-out would trade one bottleneck for a rate-limit wall, so
    the semaphore must honour the configured ceiling."""
    import asyncio

    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    in_flight = 0
    peak_in_flight = 0

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            nonlocal in_flight, peak_in_flight
            in_flight += 1
            peak_in_flight = max(peak_in_flight, in_flight)
            try:
                await asyncio.sleep(0.02)
                payload = json.loads(user.split("<user_content>")[1].split("</user_content>")[0])
                ref = payload["items"][0]["ref"]
                return json.dumps([{"ref": ref, "title": "t", "steps": []}])
            finally:
                in_flight -= 1

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    agent._logs = []
    agent.llm = _FakeLLM()

    items = [{"ref": f"r{n}", "mode": "create", "title": "t"} for n in range(1, 11)]
    result = asyncio.run(agent.run(items=items, concurrency=2))

    check("tas.refine_bounded_success", len(result.data["refined"]) == 10, str(result.data))
    check(
        "tas.refine_respects_limit",
        peak_in_flight <= 2,
        f"{peak_in_flight} calls were in flight with a limit of 2",
    )


def test_a_raw_control_character_does_not_discard_the_response():
    """Regression: a run lost TC-05 to "Invalid control character at: line 12
    column 33". The model had written a literal newline inside a string value
    instead of the \\n escape.

    Strict decoding is right for a document written by a program. A model
    writing prose into a JSON string is a different source, and the byte
    carries no ambiguity — rejecting the whole response over it discarded work
    that parses fine without strict mode.
    """
    from pydantic import BaseModel

    from app.llm.structured import parse_and_validate_llm_list

    class _Row(BaseModel):
        ref: str
        title: str

    rows = parse_and_validate_llm_list('[{"ref": "r5", "title": "line one\nline two"}]', _Row)
    check("tas.json_control_char_tolerated", len(rows) == 1, str(rows))
    check("tas.json_control_char_value_kept", "line one" in rows[0]["title"], str(rows))

    # Loosened for control characters only — a genuinely broken document must
    # still fail, or the parser would start inventing structure.
    try:
        parse_and_validate_llm_list('[{"ref": "r5", "title": "unterminated', _Row)
        check("tas.json_truncation_still_fails", False, "a cut-off document parsed")
    except ValueError as exc:
        check("tas.json_truncation_still_fails", "truncated" in str(exc).lower(), str(exc))


def test_a_call_that_never_returns_is_cut_off_and_named():
    """Nothing else bounds one call: the client timeout is httpx's per-read
    value, and the studio bypasses BaseAgent's asyncio.wait_for. One call ran
    six minutes past its last activity and held the run open at 61%.

    The cancelled batch must also explain itself — a bare TimeoutError
    stringifies to nothing, and that empty string is what the user would read
    in the run's skipped list.
    """
    import asyncio

    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            payload = json.loads(user.split("<user_content>")[1].split("</user_content>")[0])
            ref = payload["items"][0]["ref"]
            if ref == "r2":
                await asyncio.sleep(30)  # never returns within the budget
            return json.dumps([{"ref": ref, "title": "t", "steps": []}])

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    agent._logs = []
    agent.llm = _FakeLLM()

    items = [{"ref": f"r{n}", "mode": "create", "title": "t"} for n in range(1, 4)]
    started = time.monotonic()
    result = asyncio.run(agent.run(items=items, concurrency=3, call_timeout=0.2))
    elapsed = time.monotonic() - started

    check("tas.refine_timeout_bounded", elapsed < 5, f"run took {elapsed:.1f}s - the ceiling did not fire")
    check(
        "tas.refine_timeout_keeps_others",
        [row["ref"] for row in result.data["refined"]] == ["r1", "r3"],
        str([row["ref"] for row in result.data["refined"]]),
    )
    failures = result.data["failures"]
    check("tas.refine_timeout_reported", [f["ref"] for f in failures] == ["r2"], str(failures))
    check(
        "tas.refine_timeout_has_a_reason",
        "did not finish" in (failures[0]["error"] or ""),
        f"unhelpful reason: {failures[0]['error']!r}",
    )


def test_the_call_ceiling_excludes_time_spent_queued():
    """The budget is per call, not per batch-including-its-wait. With the
    ceiling around the semaphore instead of the request, a batch queued behind
    slower siblings would be failed for waiting its turn."""
    import asyncio

    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            payload = json.loads(user.split("<user_content>")[1].split("</user_content>")[0])
            ref = payload["items"][0]["ref"]
            await asyncio.sleep(0.15)
            return json.dumps([{"ref": ref, "title": "t", "steps": []}])

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    agent._logs = []
    agent.llm = _FakeLLM()

    # Six items, one slot: the last waits ~0.75s to start while each call takes
    # 0.15s. A 0.4s budget must pass all six.
    items = [{"ref": f"r{n}", "mode": "create", "title": "t"} for n in range(1, 7)]
    result = asyncio.run(agent.run(items=items, concurrency=1, call_timeout=0.4))

    check(
        "tas.refine_queue_time_not_charged",
        len(result.data["refined"]) == 6,
        f"{len(result.data['failures'])} batch(es) failed for time spent queued",
    )


def test_every_studio_llm_call_sits_under_a_wall_clock():
    """The ceiling is only worth having if no call site is left outside it.

    Read from the source, like the max_tokens guard above: the invariant is
    "no generate() call in this package is awaited bare", which is a property
    of the call sites rather than of any one execution path. The contract
    agent matters most — Playwright is the compiled framework, so it is the
    default script path, and a ceiling that skipped it would leave the busiest
    call unbounded.
    """
    import ast
    import inspect

    from app.agents.test_automation_studio import (
        contract_agent,
        coverage_agent,
        refinement_agent,
        script_agent,
    )

    for module in (contract_agent, coverage_agent, refinement_agent, script_agent):
        short = module.__name__.rsplit(".", 1)[-1]
        tree = ast.parse(inspect.getsource(module))

        # A generate() call is bounded when it is an argument to
        # with_ceiling(...) rather than the thing being awaited directly.
        bounded = set()
        for node in ast.walk(tree):
            if (
                isinstance(node, ast.Call)
                and isinstance(node.func, ast.Attribute)
                and node.func.attr == "with_ceiling"
            ):
                for inner in ast.walk(node):
                    if isinstance(inner, ast.Call) and isinstance(inner.func, ast.Attribute):
                        if inner.func.attr in {"generate", "achat", "acomplete"}:
                            bounded.add(inner.lineno)

        calls = [
            node
            for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Attribute)
            and node.func.attr in {"generate", "achat", "acomplete"}
        ]
        check(f"tas.{short}_has_calls", bool(calls), "no LLM calls found - check would pass vacuously")
        for call in calls:
            check(
                f"tas.{short}_line{call.lineno}_bounded",
                call.lineno in bounded,
                f"LLM call at line {call.lineno} in {short} is awaited with no wall clock",
            )


def test_a_timed_out_call_says_so_in_every_agent():
    """A bare TimeoutError stringifies to "", and these handlers put str(exc)
    into a toast, an agent-run error or a skipped-item reason."""
    import asyncio

    from app.agents.test_automation_studio.call_budget import LLMCallTimedOut, with_ceiling

    async def _hang():
        await asyncio.sleep(30)

    async def _drive():
        try:
            await with_ceiling(_hang(), 0.05, what="this script", setting="TAS_X_SECONDS")
        except LLMCallTimedOut as exc:
            return str(exc)
        return ""

    started = time.monotonic()
    message = asyncio.run(_drive())
    elapsed = time.monotonic() - started

    check("tas.ceiling_fires", elapsed < 5, f"took {elapsed:.1f}s")
    check("tas.ceiling_names_the_work", "this script" in message, message)
    check("tas.ceiling_names_the_setting", "TAS_X_SECONDS" in message, message)
    check("tas.ceiling_message_not_empty", len(message) > 40, repr(message))

    # A cancelled call has already spent its whole budget, so re-issuing it
    # buys another budget's worth of waiting. It must not look retriable.
    check(
        "tas.ceiling_not_a_timeout_subclass",
        not issubclass(LLMCallTimedOut, TimeoutError),
        "LLMCallTimedOut subclasses TimeoutError and would be retried as transient",
    )


def test_a_disabled_ceiling_awaits_normally():
    """Every one of these settings documents "set to 0 to disable"."""
    import asyncio

    from app.agents.test_automation_studio.call_budget import with_ceiling

    async def _quick():
        return "done"

    for budget in (0, None):
        result = asyncio.run(with_ceiling(_quick(), budget, what="x", setting="Y"))
        check(f"tas.ceiling_disabled_{budget}", result == "done", str(result))


def test_one_failing_refinement_does_not_disturb_the_calls_beside_it():
    """The whole reason for one test case per call is blast radius. Running
    them concurrently must not turn one raised exception into a lost run."""
    import asyncio

    from app.agents.test_automation_studio.refinement_agent import TestCaseRefinementAgent

    class _FakeLLM:
        async def generate(self, system, user, **kwargs):
            payload = json.loads(user.split("<user_content>")[1].split("</user_content>")[0])
            ref = payload["items"][0]["ref"]
            if ref == "r2":
                raise RuntimeError("output truncated")
            return json.dumps([{"ref": ref, "title": "t", "steps": []}])

    agent = TestCaseRefinementAgent.__new__(TestCaseRefinementAgent)
    agent._logs = []
    agent.llm = _FakeLLM()

    items = [{"ref": f"r{n}", "mode": "create", "title": "t"} for n in range(1, 5)]
    result = asyncio.run(agent.run(items=items, concurrency=4))

    check("tas.refine_partial_success", result.success is True, str(result.error))
    check(
        "tas.refine_survivors_kept",
        [row["ref"] for row in result.data["refined"]] == ["r1", "r3", "r4"],
        str([row["ref"] for row in result.data["refined"]]),
    )
    check(
        "tas.refine_failure_reported",
        [f["ref"] for f in result.data["failures"]] == ["r2"],
        str(result.data["failures"]),
    )


# ─── Script generation guards ────────────────────────────────────────────────

def test_script_agent_rejects_unknown_frameworks():
    import asyncio

    from app.agents.test_automation_studio.script_agent import (
        SUPPORTED_FRAMEWORKS,
        ScriptGenerationAgent,
    )

    check(
        "tas.frameworks_three",
        set(SUPPORTED_FRAMEWORKS) == {"playwright", "katalon", "appium"},
        str(SUPPORTED_FRAMEWORKS),
    )
    agent = ScriptGenerationAgent.__new__(ScriptGenerationAgent)
    agent._logs = []
    result = asyncio.run(agent.run(test_case={"tc_display_id": "TC-1"}, framework="selenium"))
    check("tas.script_unknown_framework_fails", result.success is False)
    check("tas.script_unknown_framework_message", "Unsupported framework" in (result.error or ""))


def test_script_filenames_cannot_escape_the_zip():
    from app.services.test_automation_studio.script_lab_service import _safe_slug

    check("tas.slug_no_traversal", ".." not in _safe_slug("../../etc/passwd"))
    check("tas.slug_no_separator", "/" not in _safe_slug("a/b/c"))
    check("tas.slug_not_empty", _safe_slug("...") != "")


# ─── Export split (requirement 2d) ───────────────────────────────────────────

def _empty_context():
    """Export context with nothing to inherit — a purely gap-derived download."""
    return export_service._Context({}, {}, {})


def test_export_separates_manual_from_automation():
    rows = [
        _test_case(tc_display_id="TC-0001", classification="automation"),
        _test_case(tc_display_id="TC-0002", classification="manual"),
        _test_case(tc_display_id="TC-0003", classification="undecided"),
    ]
    context = _empty_context()
    workbook = export_service.to_excel(rows, context)
    check("tas.export_xlsx_produced", isinstance(workbook, bytes) and len(workbook) > 0)

    automation_csv = export_service.to_csv(
        [row for row in rows if row.classification == "automation"], context, automation=True
    ).decode("utf-8-sig")
    manual_csv = export_service.to_csv(
        [row for row in rows if row.classification == "manual"], context, automation=False
    ).decode("utf-8-sig")

    check("tas.export_automation_has_row", "TC-0001" in automation_csv)
    check("tas.export_automation_excludes_manual", "TC-0002" not in automation_csv)
    check("tas.export_manual_has_row", "TC-0002" in manual_csv)
    check("tas.export_manual_excludes_automation", "TC-0001" not in manual_csv)
    check(
        "tas.export_both_have_test_data_column",
        "Test Data Required" in automation_csv and "Test Data Required" in manual_csv,
    )


def test_download_uses_the_same_columns_as_the_upload_template():
    """The download is handed back to the team who supplied the sheet.

    Manual and Automation used to carry different, studio-invented column sets,
    so neither could be opened next to the file that was uploaded. Both now
    lead with the platform's canonical template, in order, and append the
    studio's own columns after it.
    """
    from app.services.test_case_template import TEST_CASE_TEMPLATE_HEADERS

    context = _empty_context()
    automation_csv = export_service.to_csv(
        [_test_case(classification="automation")], context, automation=True
    ).decode("utf-8-sig")
    manual_csv = export_service.to_csv(
        [_test_case(classification="manual")], context, automation=False
    ).decode("utf-8-sig")

    automation_header = automation_csv.splitlines()[0].split(",")
    manual_header = manual_csv.splitlines()[0].split(",")

    check(
        "tas.export_same_shape_both_ways",
        automation_header == manual_header,
        "Manual and Automation must be the same document shape",
    )
    check(
        "tas.export_leads_with_template",
        automation_header[: len(TEST_CASE_TEMPLATE_HEADERS)] == TEST_CASE_TEMPLATE_HEADERS,
        f"first columns were {automation_header[:6]}",
    )
    # The studio's additions come after, never interleaved, so a reader diffing
    # against their upload does not find their own columns shifted sideways.
    check(
        "tas.export_appends_studio_columns",
        automation_header[len(TEST_CASE_TEMPLATE_HEADERS) :] == export_service.STUDIO_COLUMNS,
        str(automation_header[len(TEST_CASE_TEMPLATE_HEADERS) :]),
    )
    # Steps must render in the format the importer parses back, or a download
    # cannot be re-uploaded without the steps collapsing into prose.
    check("tas.export_steps_round_trip", " -> " in automation_csv or "1. " in automation_csv)


def test_uploaded_columns_the_studio_does_not_own_are_carried_through():
    """Domain, Channel, Product and the execution columns come from the source.

    The studio rewrites the objective, steps and expected result. It has no
    opinion on the rest, and inventing values there would put data in a team's
    sheet that nobody chose — so they are copied from the row the test case
    came from, and left blank when there was none.
    """
    from app.models.test_automation_studio import TasSourceTestCase

    source = TasSourceTestCase(
        id=1,
        project_id=1,
        batch_id=1,
        tc_display_id="TC-01",
        title="Original name",
        steps=[],
        created_by=1,
        source_row={
            "domain": "Digital Channels",
            "channel": "Web",
            "product": "eSIM",
            "area_of_test": "Checkout",
            "environment": "SIT",
            "sub_request_type": "Enhancement",
            "tested_by": "A Tester",
            "jira_or_ppm": "QA-1234",
            "overall_status": "Passed",
            "test_case_complexity": "High",
        },
    )
    refined = _test_case(
        tc_display_id="TC-01", origin="imported", classification="automation"
    )
    refined.source_uploaded_test_case_id = 1

    context = export_service._Context({1: source}, {}, {})
    csv_text = export_service.to_csv([refined], context, automation=True).decode("utf-8-sig")

    for value in (
        "Digital Channels",
        "Web",
        "eSIM",
        "Checkout",
        "SIT",
        "Enhancement",
        "A Tester",
        "QA-1234",
        "Passed",
        "High",
    ):
        check(f"tas.export_carries_{value.replace(' ', '_')}", value in csv_text, csv_text[:300])

    # And the preserved ID is what identifies the row, not a studio-minted one.
    check("tas.export_keeps_uploaded_id", "TC-01" in csv_text)

    # With no source row the columns stay empty rather than being invented.
    bare = export_service.to_csv(
        [_test_case(classification="manual")], _empty_context(), automation=False
    ).decode("utf-8-sig")
    check("tas.export_invents_nothing", "Digital Channels" not in bare)


def test_download_reuses_the_uploaded_workbook_header_colours():
    """The colours carry meaning, so a uniform download loses information.

    This template tints each column by who owns it — green for what QA
    authors, salmon for what comes from the request, grey for execution
    results — and a reader scans the sheet by those bands. The export used one
    colour for every column, so the returned sheet no longer read like the one
    that was uploaded.
    """
    import io

    from openpyxl import Workbook, load_workbook

    from app.services.test_automation_studio import sheet_import

    # A workbook styled the way the source sheets are.
    source = Workbook()
    sheet = source.active
    tinted = {"ID": "FF92D050", "Domain": "FFDA9694", "Tested By": "FFA6A6A6"}
    headers = ["ID", "Domain", "Test Case ID", "Tested By"]
    sheet.append(headers)
    for index, header in enumerate(headers, start=1):
        if header in tinted:
            from openpyxl.styles import Font, PatternFill

            sheet.cell(row=1, column=index).fill = PatternFill(
                fill_type="solid", fgColor=tinted[header]
            )
            sheet.cell(row=1, column=index).font = Font(bold=True, color="FFFFFFFF", size=12)
    buffer = io.BytesIO()
    source.save(buffer)

    styles = sheet_import.read_header_format(buffer.getvalue(), "source.xlsx")
    check("tas.format_read_per_column", len(styles) >= 3, str(sorted(styles)))
    check("tas.format_keeps_green", styles.get("id", {}).get("fill") == "FF92D050", str(styles.get("id")))
    check(
        "tas.format_keeps_salmon",
        styles.get("domain", {}).get("fill") == "FFDA9694",
        str(styles.get("domain")),
    )
    check(
        "tas.format_keeps_grey",
        styles.get("tested_by", {}).get("fill") == "FFA6A6A6",
        str(styles.get("tested_by")),
    )

    # And the export replays them onto the same columns.
    context = export_service._Context({}, {}, {}, styles)
    workbook = load_workbook(
        io.BytesIO(export_service.to_excel([_test_case(classification="manual")], context))
    )
    exported = workbook["Manual Test Cases"]
    by_header = {
        exported.cell(row=1, column=col).value: exported.cell(row=1, column=col).fill.fgColor.rgb
        for col in range(1, exported.max_column + 1)
    }
    check("tas.export_colour_id", by_header.get("ID") == "FF92D050", str(by_header.get("ID")))
    check("tas.export_colour_domain", by_header.get("Domain") == "FFDA9694", str(by_header.get("Domain")))
    check(
        "tas.export_colour_tested_by",
        by_header.get("Tested By") == "FFA6A6A6",
        str(by_header.get("Tested By")),
    )
    # A column the source never styled still has to look like a header.
    check("tas.export_colour_default", bool(by_header.get("Priority")), str(by_header.get("Priority")))

    # A CSV source has no styling to read, and that must not fail the export.
    check("tas.format_csv_is_empty", sheet_import.read_header_format(b"a,b\n1,2\n", "x.csv") == {})


def test_test_data_required_column_renders_yes_no():
    csv_text = export_service.to_csv(
        [_test_case(test_data_required=True, test_data_status="needs_user_action")],
        _empty_context(),
        automation=False,
    ).decode("utf-8-sig")
    check("tas.export_yes", ",Yes," in csv_text or csv_text.rstrip().endswith("Yes"), csv_text[:400])


# ─── URL validation ──────────────────────────────────────────────────────────

def test_application_url_rejects_non_http_schemes():
    from app.schemas.test_automation_studio import normalize_application_url

    check("tas.url_https_ok", normalize_application_url("https://app.test") == "https://app.test")
    check("tas.url_blank_none", normalize_application_url("   ") is None)
    for bad in ("javascript:alert(1)", "file:///etc/passwd", "app.test"):
        try:
            normalize_application_url(bad)
            check(f"tas.url_rejects_{bad[:12]}", False, f"{bad} was accepted")
        except ValueError:
            check(f"tas.url_rejects_{bad[:12]}", True)


# ─── Deleting generated artefacts ────────────────────────────────────────────

DELETE_ROUTES = {
    # path -> the permission constant that also produces the artefact
    "/projects/{project_id}/requirements/delete": "TAS_ASSESS_COVERAGE",
    "/projects/{project_id}/source-test-cases/delete": "TAS_INTAKE",
    "/projects/{project_id}/test-cases/delete": "TAS_REFINE_TEST_CASES",
    "/projects/{project_id}/scripts/delete": "TAS_GENERATE_SCRIPT",
}


def test_every_generated_artefact_can_be_deleted():
    """All four artefact families the studio generates must have a delete.

    Without one, the only way out of a bad generation run is a DBA: the
    screens regenerate into new versions rather than replacing, so mistakes
    accumulate.
    """
    from app.main import app

    paths = app.openapi()["paths"]
    for path in DELETE_ROUTES:
        full = f"/api/v1/lab/test-automation-studio{path}"
        check(f"tas.delete_route_exists:{path}", full in paths, f"{full} missing")


def test_deleting_an_artefact_needs_the_permission_that_generated_it():
    """Delete is gated as tightly as generation, and never on `tas.view`.

    Binding each delete to the permission that produces that artefact is what
    keeps a role admitted to the studio to read or approve from emptying the
    workspace.
    """
    import inspect

    from app.api.v1.endpoints import test_automation_studio as tas_endpoints

    by_path = {route.path: route for route in tas_endpoints.router.routes}
    for path, permission in DELETE_ROUTES.items():
        route = by_path.get(path)
        check(f"tas.delete_route_registered:{path}", route is not None)
        if route is None:
            continue
        source = inspect.getsource(route.endpoint)
        check(
            f"tas.delete_guarded_by:{path}",
            f"_guard({permission}" in source,
            f"{path} is not guarded by {permission}",
        )

    # And no read-only or execution-only role may hold any of them.
    for role in ("Viewer/Auditor", "Tester", "Business Analyst"):
        held = rbac_service.ROLE_PERMISSIONS[role]
        for permission in sorted(set(DELETE_ROUTES.values())):
            check(
                f"tas.{role}_cannot_delete:{permission}",
                getattr(rbac_service, permission) not in held,
                f"{role} holds {permission} - it could delete generated artefacts",
            )


def test_deleting_a_test_case_takes_its_scripts_with_it():
    """The cascade the delete relies on lives on the foreign key.

    `delete_test_cases` issues a Core DELETE (an ORM cascade would have to
    lazy-load `scripts` mid-flush, which raises under asyncio), so the
    database is what actually removes the scripts. If this FK ever loses its
    ON DELETE CASCADE the delete starts failing on a constraint violation
    instead.
    """
    from app.models.test_automation_studio import TasScriptAsset

    column = TasScriptAsset.__table__.c.refined_test_case_id
    foreign_key = next(iter(column.foreign_keys))
    check(
        "tas.script_fk_cascades",
        foreign_key.ondelete == "CASCADE",
        f"scripts FK ondelete is {foreign_key.ondelete!r} - a test case delete would be blocked",
    )
    check("tas.script_fk_not_nullable", column.nullable is False)


def test_delete_services_exist_for_every_artefact():
    from app.services.test_automation_studio import (
        coverage_service,
        refinement_service,
        script_lab_service,
    )

    for fn, name in (
        (coverage_service.delete_requirements, "delete_requirements"),
        (coverage_service.delete_source_test_cases, "delete_source_test_cases"),
        (refinement_service.delete_test_cases, "delete_test_cases"),
        (script_lab_service.delete_scripts, "delete_scripts"),
    ):
        check(f"tas.{name}_callable", callable(fn))


def test_a_delete_reports_what_else_it_removed():
    """The summary must carry the collateral, not just a count.

    Deleting a refined test case also removes superseded versions and
    cascades scripts. The screens tell the user that after the fact, so the
    contract has to have somewhere to say it.
    """
    from app.schemas.test_automation_studio import BulkDeleteRequest, DeletionSummary

    summary = DeletionSummary()
    for field in (
        "deleted",
        "not_found",
        "versions_deleted",
        "scripts_deleted",
        "test_cases_unlinked",
        "approved_deleted",
    ):
        check(f"tas.deletion_summary_has_{field}", hasattr(summary, field))

    # An empty request would delete nothing and is almost always a bug in the
    # caller, so it is refused at the contract.
    try:
        BulkDeleteRequest(ids=[])
        check("tas.delete_rejects_empty_ids", False, "an empty id list was accepted")
    except Exception:
        check("tas.delete_rejects_empty_ids", True)


# ─── Losing the source a test case was refined from ──────────────────────────

def test_a_refined_test_case_reports_when_its_source_is_gone():
    """Both source links are SET NULL, so deletion strands the refined row.

    Observed live: five refined test cases sat in a project with every source
    link null. The grid showed them as ordinary rows, and because the
    "already refined" check reads that same nulled column they no longer
    counted as done — the next run would have built a second test case for
    each. The state is legitimate and not worth blocking, but it has to be
    visible.
    """
    from app.schemas.test_automation_studio import RefinedTestCaseOut

    def _out(**overrides) -> RefinedTestCaseOut:
        payload = dict(
            id=1,
            project_id=5,
            origin="imported",
            tc_display_id="TC-01",
            title="Validate legacy redirect",
            priority="Medium",
            classification="automation",
            test_data_required=False,
            test_data_status="not_required",
            status="approved",
            version=1,
            is_current=True,
            edited_by_user=False,
            created_at=datetime(2026, 8, 11, tzinfo=timezone.utc),
            updated_at=datetime(2026, 8, 11, tzinfo=timezone.utc),
        )
        payload.update(overrides)
        return RefinedTestCaseOut(**payload)

    check(
        "tas.source_missing_when_uploaded_row_deleted",
        _out(source_uploaded_test_case_id=None, source_test_case_id=None).source_missing,
        "an imported test case with no source link is not reported as orphaned",
    )
    check(
        "tas.source_present_is_not_flagged",
        not _out(source_uploaded_test_case_id=55).source_missing,
        "a linked test case was wrongly flagged as orphaned",
    )
    check(
        "tas.platform_source_present_is_not_flagged",
        not _out(origin="existing", source_test_case_id=12).source_missing,
        "a test case linked to a platform row was wrongly flagged",
    )
    # A gap-derived test case never had a source, so it is not orphaned.
    check(
        "tas.derived_is_never_orphaned",
        not _out(origin="derived").source_missing,
        "a gap-derived test case was reported as having lost a source",
    )
    check(
        "tas.source_missing_is_serialised",
        "source_missing" in _out().model_dump(),
        "source_missing is computed but never reaches the client",
    )


def test_refinement_recognises_an_orphaned_test_case_by_its_id():
    """The duplicate guard cannot rest on the FK alone.

    Re-extracting a document mints new source rows. If the refined test case
    for that ID lost its link when the old rows went, matching only on
    `source_uploaded_test_case_id` makes every row look unrefined and the run
    produces a second test case per behaviour.
    """
    import inspect

    from app.services.test_automation_studio import refinement_service

    source = inspect.getsource(refinement_service.generate_refined_test_cases)
    check(
        "tas.dedupe_matches_display_id",
        "already_display_ids" in source,
        "generation no longer falls back to the display ID - orphans will duplicate",
    )


def test_an_unmatched_agent_row_is_reported_not_dropped():
    """A returned row nobody asked for must not vanish in silence.

    The loop that persists the agent's output skips any row whose `source_ref`
    does not resolve. Continuing without a word would leave the run reporting
    a clean success while the grid is short an item and nothing anywhere says
    why.
    """
    import inspect

    from app.services.test_automation_studio import refinement_service

    source = inspect.getsource(refinement_service.generate_refined_test_cases)
    marker = "entry = context.get(str(row.get(\"source_ref\")))"
    check("tas.persist_loop_found", marker in source, "the persistence loop moved")
    if marker in source:
        tail = source.split(marker, 1)[1]
        # Whatever the branch does, it must record something before moving on.
        branch = tail.split("continue", 1)[0]
        check(
            "tas.unmatched_row_is_recorded",
            "skipped.append" in branch,
            "an unresolvable agent row is still dropped without a skipped entry",
        )


# Functions only, and only ones defined here: `test_data_bridge` is an
# imported module whose name also starts with "test_".
TESTS = [
    value
    for name, value in list(globals().items())
    if name.startswith("test_")
    and callable(value)
    and getattr(value, "__module__", None) == __name__
    and name != "test_all"
]


def test_all():
    """Single pytest entry point so the harness reports one result per file."""
    for fn in TESTS:
        _run(fn)
    assert not _FAIL, "\n".join(f"{name}: {detail}" for name, detail in _FAIL)


if __name__ == "__main__":
    for fn in TESTS:
        _run(fn)
    print(f"{len(_PASS)} passed, {len(_FAIL)} failed")
    for name, detail in _FAIL:
        print(f"  FAIL {name}: {detail}")
    sys.exit(1 if _FAIL else 0)
